import streamlit as st
import pandas as pd
import re
from collections import defaultdict

st.set_page_config(page_title="HCP Qual Insight Engine", layout="wide", page_icon="🔬")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=Geist:wght@300;400;500;600&display=swap');
html,body,[data-testid="stAppViewContainer"]{background:#080c14!important;color:#c8d4e8;font-family:'Geist',sans-serif}
[data-testid="stHeader"]{background:transparent!important}
.block-container{padding:1.5rem 2rem!important;max-width:1600px!important}
.stTextInput>div>div>input{background:#0d1420!important;border:1px solid #1e2d47!important;border-radius:10px!important;color:#e2ecf8!important;font-size:15px!important;padding:12px 18px!important}
.card{background:#0d1420;border:1px solid #1a2640;border-radius:12px;padding:18px 22px;margin-bottom:14px}
.stat-num{font-family:'IBM Plex Mono',monospace;font-size:38px;font-weight:600;color:#e2ecf8;line-height:1}
.stat-lbl{font-size:10px;color:#4a6080;letter-spacing:1.5px;text-transform:uppercase;margin-top:4px}
.sec-lbl{font-size:10px;letter-spacing:2px;text-transform:uppercase;color:#3b6ef7;font-weight:600;margin-bottom:12px;font-family:'IBM Plex Mono',monospace}
.fq-card{background:#0a1220;border:1px solid #1a2640;border-left:3px solid #3b6ef7;border-radius:0 12px 12px 0;padding:16px 20px;margin-bottom:12px}
.fq-card.comm{border-left-color:#f59e0b}
.fq-card.acad{border-left-color:#3b6ef7}
.fq-meta{font-family:'IBM Plex Mono',monospace;font-size:10px;color:#4a6080;margin-bottom:8px}
.fq-text{font-size:13px;color:#c8d4e8;line-height:1.9;font-style:italic}
.hl-ins{background:rgba(249,115,22,.25);border-radius:3px;padding:1px 4px;color:#fb923c;font-style:normal;font-weight:600}
.hl-pfs{background:rgba(59,130,246,.25);border-radius:3px;padding:1px 4px;color:#60a5fa;font-style:normal;font-weight:600}
.hl-os{background:rgba(139,92,246,.25);border-radius:3px;padding:1px 4px;color:#a78bfa;font-style:normal;font-weight:600}
.hl-saf{background:rgba(245,158,11,.25);border-radius:3px;padding:1px 4px;color:#fbbf24;font-style:normal;font-weight:600}
.hl-sz{background:rgba(239,68,68,.25);border-radius:3px;padding:1px 4px;color:#f87171;font-style:normal;font-weight:600}
.hl-qol{background:rgba(16,185,129,.25);border-radius:3px;padding:1px 4px;color:#34d399;font-style:normal;font-weight:600}
.hl-oral{background:rgba(132,204,22,.25);border-radius:3px;padding:1px 4px;color:#a3e635;font-style:normal;font-weight:600}
.hl-trial{background:rgba(251,191,36,.25);border-radius:3px;padding:1px 4px;color:#fcd34d;font-style:normal;font-weight:600}
.hl-nccn{background:rgba(167,139,250,.25);border-radius:3px;padding:1px 4px;color:#c4b5fd;font-style:normal;font-weight:600}
.hl-def{background:rgba(59,110,247,.2);border-radius:3px;padding:1px 4px;color:#93c5fd;font-style:normal;font-weight:600}
.tag{display:inline-block;padding:2px 8px;border-radius:99px;font-size:10px;font-weight:600;margin-right:4px;margin-top:2px}
.bar-row{display:flex;align-items:center;gap:10px;margin-bottom:9px}
.bar-lbl{font-size:11px;color:#8a9ab5;min-width:180px}
.bar-trk{flex:1;height:5px;background:#1a2640;border-radius:3px}
.bar-fill{height:5px;border-radius:3px}
.bar-cnt{font-family:'IBM Plex Mono',monospace;font-size:11px;color:#4a6080;min-width:60px;text-align:right}
.diff-pos{color:#10b981;font-weight:700}
.diff-neg{color:#ef4444;font-weight:700}
.seg-a{background:rgba(245,158,11,.15);color:#fbbf24;border:1px solid rgba(245,158,11,.3);padding:2px 10px;border-radius:99px;font-size:11px;font-weight:600}
.seg-b{background:rgba(59,110,247,.15);color:#93c5fd;border:1px solid rgba(59,110,247,.3);padding:2px 10px;border-radius:99px;font-size:11px;font-weight:600}
.ibadge{display:inline-block;padding:3px 12px;border-radius:99px;font-size:10px;font-weight:600;letter-spacing:1px;text-transform:uppercase;background:rgba(59,110,247,.12);color:#3b6ef7;border:1px solid rgba(59,110,247,.25);margin-right:6px}
.tbadge{display:inline-block;padding:3px 12px;border-radius:99px;font-size:10px;background:rgba(16,185,129,.1);color:#34d399;border:1px solid rgba(16,185,129,.2);margin-right:6px}
[data-testid="stExpander"]{background:#0d1420!important;border:1px solid #1a2640!important;border-radius:10px!important}
.stButton>button{background:#0d1420!important;border:1px solid #1e2d47!important;color:#8a9ab5!important;border-radius:8px!important;font-size:11px!important;padding:5px 10px!important;width:100%!important;text-align:left!important;white-space:normal!important;height:auto!important}
.stButton>button:hover{border-color:#3b6ef7!important;color:#93c5fd!important}
.stSelectbox>div>div{background:#0d1420!important;border:1px solid #1e2d47!important;color:#c8d4e8!important;border-radius:8px!important}
hr{border:none;border-top:1px solid #1a2640;margin:20px 0}
</style>
""", unsafe_allow_html=True)

# ── THEMES ────────────────────────────────────────────────────────────────────
THEMES = {
    "PFS":                 ["progression free survival","progression-free","pfs","time to progression","time to next intervention","time to next treatment"],
    "OS":                  ["overall survival"," os ","lifespan","live longer","living longer","survivability"],
    "Safety/Tolerability": ["side effect","adverse event","toxicity","tolerab","tolerat","fatigue","nausea","hepato","liver function","myelosuppression","hematotoxicity","well tolerated"],
    "Seizures":            ["seizure","epilepsy","seizure control","seizure frequency","seizure reduction","seizure prophylaxis","anti-seizure"],
    "Quality of Life":     ["quality of life","qol","daily activities","activities of daily","functional","cognitive","cognition","brain fog","neurocognit"],
    "Efficacy":            ["efficacy","effective","response rate","tumor reduction","tumor shrinkage","disease control","proven benefit","superior efficacy"],
    "Oral Administration": ["oral","pill","tablet","once daily","easy to take","easy to administer","oral formulation","oral daily"],
    "Cost/Insurance":      ["reimburs","insurance","prior auth","formulary","coverage","copay","co-pay","cost","afford","appeal","denial","out of pocket","financial","free drug"],
    "Patient Assistance":  ["patient assistance","patient support","financial assistance","manufacturer assistance","copay program","compassionate use","patient program"],
    "NCCN/Guidelines":     ["nccn","guideline","standard of care","fda approved","fda approval","category 1","category 2","institutional","formulary"],
    "Indigo Trial":        ["indigo trial","indigo study","phase 3","phase three","clinical trial data","indigo","nejm","new england journal"],
    "Radiation Delay":     ["delay radiation","defer radiation","avoid radiation","delay radiotherapy","delay chemo","defer chemo","avoid chemo"],
    "Surgery/Resection":   ["resection","biopsy","subtotal resection","gross total resection","residual tumor","residual disease"],
    "Targeted Therapy":    ["targeted therapy","targeted","mechanism of action","isocitrate","targeted treatment","idh inhibitor","blood-brain barrier"],
    "IDH Mutation":        ["idh1","idh2","idh mutation","idh mutant","idh inhibitor","isocitrate dehydrogenase","idh status"],
    "Fertility":           ["fertilit","birth defect","pregnan","desire to have children","family planning"],
    "Performance Status":  ["ecog","performance status","kps","fit patient","younger patient","functional status"],
    "Rep Support":         ["representative","drug rep","pharma rep"," rep ","lunch","dinner","detail","rep support"],
}

TC = {
    "PFS":"#3b82f6","OS":"#8b5cf6","Safety/Tolerability":"#f59e0b",
    "Seizures":"#ef4444","Quality of Life":"#10b981","Efficacy":"#06b6d4",
    "Oral Administration":"#84cc16","Cost/Insurance":"#f97316",
    "Patient Assistance":"#ec4899","NCCN/Guidelines":"#a78bfa",
    "Indigo Trial":"#fbbf24","Radiation Delay":"#fb7185",
    "Surgery/Resection":"#94a3b8","Targeted Therapy":"#34d399",
    "IDH Mutation":"#22d3ee","Fertility":"#f472b6",
    "Performance Status":"#60a5fa","Rep Support":"#fcd34d",
}

HL_MAP = {
    "Cost/Insurance":"hl-ins","Patient Assistance":"hl-ins",
    "PFS":"hl-pfs","OS":"hl-os","Safety/Tolerability":"hl-saf","Seizures":"hl-sz",
    "Quality of Life":"hl-qol","Oral Administration":"hl-oral",
    "Indigo Trial":"hl-trial","NCCN/Guidelines":"hl-nccn",
}

DRIVER_MAP = {
    "Efficacy / PFS Data":     ["efficacy","effective","response","pfs","progression free","survival benefit","indigo","phase 3"],
    "Tolerability":            ["tolerab","tolerat","side effect","well tolerated","low toxicity","manageable","minimal side"],
    "Oral Convenience":        ["oral","pill","tablet","once daily","convenient","easy to take"],
    "Cost / Access":           ["cost","afford","insurance","reimburs","copay","access","patient assistance","denied"],
    "Targeted Mechanism":      ["targeted","idh","mechanism of action","targeted therapy","isocitrate","blood-brain"],
    "Delay Radiation/Chemo":   ["delay radiation","defer radiation","avoid radiation","delay chemo","defer chemo"],
    "Clinical Trial (Indigo)": ["indigo","phase 3","clinical trial","nejm","new england journal"],
    "NCCN / FDA":              ["nccn","guideline","standard of care","fda"],
    "Seizure Control":         ["seizure","epilepsy","seizure control","seizure frequency"],
    "Quality of Life":         ["quality of life","qol","daily activities","cognitive","neurocognit"],
    "Patient Selection":       ["grade 2","idh mutant","residual tumor","subtotal resection","mutation status","performance status"],
}

# ── EXCEL LOADER ──────────────────────────────────────────────────────────────
def load_excel(file):
    """
    Robust loader handling:
    1. Sagan_Drivers format — header at row 0: Target | Specialty | Setting | Text
    2. Sagan_Responses format — metadata rows at top, data starts at row 8+
    3. Generic qual format — auto-detect
    """
    import openpyxl

    # ── Step 1: scan with openpyxl to find the real header row ───────────────
    try:
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return None, f"Cannot open file: {e}"

    if not all_rows:
        return None, "File is empty."

    # Find first row with 3+ non-null values — that's our header/data start
    header_row = 0
    for i, row in enumerate(all_rows):
        non_null = [v for v in row if v is not None]
        if len(non_null) >= 3:
            header_row = i
            break

    # ── Step 2: read with pandas using correct header row ─────────────────────
    try:
        raw = pd.read_excel(file, header=header_row)
    except Exception as e:
        return None, f"Cannot read file: {e}"

    if raw.empty:
        return None, "No data found after header row."

    # Convert all to string for uniform processing
    raw = raw.astype(str).replace("nan", pd.NA).replace("None", pd.NA)

    # ── Step 3: identify text column (highest avg string length) ─────────────
    avg_lens = {}
    for c in raw.columns:
        lengths = raw[c].dropna().apply(lambda x: len(str(x)))
        avg_lens[c] = lengths.mean() if len(lengths) > 0 else 0

    if not avg_lens or max(avg_lens.values()) < 30:
        return None, "No verbatim text column found. Check your file has a column with interview responses."

    text_col = max(avg_lens, key=avg_lens.get)

    # ── Step 4: filter to real content rows ───────────────────────────────────
    mask = raw[text_col].apply(lambda x: isinstance(x, str) and len(x.strip()) > 80)
    data = raw[mask].reset_index(drop=True)

    if len(data) == 0:
        return None, f"No rows with text >80 chars in column '{text_col}'. Check your verbatim column."

    # ── Step 5: identify metadata columns ─────────────────────────────────────
    meta_cols = []
    for c in raw.columns:
        if c == text_col: continue
        col_data = data[c].dropna()
        if len(col_data) == 0: continue
        avg_len = col_data.apply(lambda x: len(str(x))).mean()
        n_unique = col_data.nunique()
        if avg_len < 60 and 1 < n_unique <= 30:
            meta_cols.append(c)

    def best_col(keywords, fallback="Unknown"):
        """Match by column values first, then column header name."""
        # Try matching keywords against column VALUES
        for c in meta_cols:
            vals = data[c].dropna().astype(str).str.upper().tolist()
            if any(any(kw.upper() in v for kw in keywords) for v in vals):
                return data[c].fillna(fallback).astype(str).str.strip()
        # Try matching keywords against COLUMN NAMES
        for c in meta_cols:
            col_name = str(c).upper()
            if any(kw.upper() in col_name for kw in keywords):
                return data[c].fillna(fallback).astype(str).str.strip()
        # Fallback: first available meta col
        if meta_cols:
            return data[meta_cols[0]].fillna(fallback).astype(str).str.strip()
        return pd.Series([fallback] * len(data))

    out = pd.DataFrame()
    out["id"]        = [f"R_{i+1:03d}" for i in range(len(data))]
    out["text"]      = data[text_col].astype(str)

    # Setting — look for Academic/Community in values OR "setting/practice" in header
    out["setting"]   = best_col(["COMMUNITY","ACADEMIC","PRIVATE","HOSPITAL","TEACHING",
                                  "SETTING","PRACTICE","SITE","INSTITUTION"])
    # Simplify setting values: extract Community/Academic/Private from long strings
    def simplify_setting(v):
        vu = str(v).upper()
        if "ACADEMIC" in vu: return "Academic"
        if "COMMUNITY" in vu: return "Community"
        if "PRIVATE" in vu: return "Private Practice"
        if "TEACHING" in vu: return "Teaching Hospital"
        return v
    out["setting"] = out["setting"].apply(simplify_setting)

    out["target"]    = best_col(["TL","TARGET","TIER","ON TL","OFF TL","CO-LOC",
                                  "COLOC","PRIORITY","LIST"])
    out["specialty"] = best_col(["ONCOL","ONCOLOGY","NEURO","HEMATOL","MEDONC",
                                  "HEMEONC","SPEC","DISCIPLINE","HCP TYPE","NEUROLOGY",
                                  "CLINICAL","MEDICAL"])
    # Simplify specialty values
    def simplify_specialty(v):
        vu = str(v).upper()
        if "NEURO-ONC" in vu or "NEURO ONC" in vu: return "Neuro-Onc"
        if "NEUROLOGY" in vu: return "Neurology"
        if "HEMATOL" in vu: return "HemeOnc"
        if "MEDICAL" in vu or "CLINICAL" in vu: return "MedOnc"
        return v
    out["specialty"] = out["specialty"].apply(simplify_specialty)

    # ── Step 6: extract doctor-only speech ────────────────────────────────────
    def doctor_text(t):
        """
        Extract ONLY doctor/HCP speech from AI-moderated transcripts.
        Steps:
          1. Detect which speaker label (A or B) is the moderator vs doctor
             by scoring against moderator phrase patterns.
          2. Extract doctor chunks only.
          3. Filter out any chunk that is purely a question (ends in ?)
             or is a known moderator-leak pattern (repeating the question back).
          4. Clean and join.
        """
        MOD_PHRASES = [
            'thank you', 'could you', 'when you', 'what specific', 'how do you',
            'to start,', 'are you still there', 'i am here to understand',
            "i'm here to understand", 'hello doctor', 'hello,', 'hello.',
            'that is helpful', 'great,', 'understood.', 'i see,', 'interesting,',
            'can you elaborate', 'could you clarify', 'could you describe',
            'what are the primary', 'what factors', 'how does', 'how would you',
            'in your experience', 'to confirm', 'so to confirm', 'you mentioned',
            'i did not quite catch', 'please clarify', 'tell me more',
            'how do i talk', 'how do you navigate', 'how do you handle',
        ]

        # Patterns that indicate a chunk is the moderator question being echoed back
        MOD_ECHO_PATTERNS = [
            r'^so when \w+ or \w+ becomes a',
            r'^when \w+ or \w+ coverage becomes',
            r'^how do (i|you) (talk|navigate|handle|address)',
            r'^is that your question',
            r'^can (you|we) (repeat|go to|move)',
            r'^(repeat|can you repeat)',
            r'^(hello|hi|bye|goodbye)[\.\?]?$',
            r'^yes[\.\?]?$',
            r'^no[\.\?]?$',
            r'^(okay|ok|alright|sure)[\.\,]?$',
        ]

        def clean_chunk(chunk):
            """Return cleaned chunk or None if it should be excluded."""
            c = re.sub(r'\s+', ' ', chunk).strip()
            if len(c) < 15:
                return None
            cl = c.lower()
            # Exclude pure questions (moderator asking) — BUT keep answers that
            # start with context and end in ? (e.g. 'IDH mutation. Yes?')
            # Rule: if >60% of sentences end in ?, exclude the chunk
            sentences = [s.strip() for s in re.split(r'(?<=[.!?])\s+', c) if s.strip()]
            q_count = sum(1 for s in sentences if s.endswith('?'))
            if sentences and q_count / len(sentences) > 0.6:
                return None
            # Exclude moderator echo patterns
            for pat in MOD_ECHO_PATTERNS:
                if re.match(pat, cl):
                    return None
            # Exclude very short filler responses
            filler = ['you.','me.','yes.','no.','okay.','ok.','sure.','hello.','bye.','goodbye.','repeat the question','repeat your question','can you repeat','i cannot hear','i could not hear']
            if any(cl.startswith(f) or cl == f.rstrip('.') for f in filler):
                return None
            return c

        # ── Extract speaker chunks ────────────────────────────────────────
        b_chunks = re.findall(r'SPEAKER_B:\s*(.+?)(?=SPEAKER_A:|SPEAKER_B:|\Z)', t, re.DOTALL)
        a_chunks = re.findall(r'SPEAKER_A:\s*(.+?)(?=SPEAKER_A:|SPEAKER_B:|\Z)', t, re.DOTALL)

        if a_chunks or b_chunks:
            b_text = ' '.join(b_chunks).lower()
            a_text = ' '.join(a_chunks).lower()
            b_mod_score = sum(1 for p in MOD_PHRASES if p in b_text)
            a_mod_score = sum(1 for p in MOD_PHRASES if p in a_text)

            doctor_chunks = a_chunks if b_mod_score > a_mod_score else b_chunks
            cleaned = [clean_chunk(c) for c in doctor_chunks]
            cleaned = [c for c in cleaned if c]
            result = ' '.join(cleaned)
            if len(result) > 30:
                return result

        # Doctor: / HCP: label patterns (alternative transcript formats)
        for pattern in [
            r'Doctor:\s*(.+?)(?=AI Moderator:|Doctor:|\Z)',
            r'HCP:\s*(.+?)(?=Moderator:|HCP:|\Z)',
        ]:
            lines = re.findall(pattern, t, re.DOTALL)
            cleaned = [clean_chunk(l) for l in lines]
            cleaned = [c for c in cleaned if c]
            result = ' '.join(cleaned)
            if len(result) > 30:
                return result

        return re.sub(r'\s+', ' ', t).strip()

    out["text"]       = out["text"].apply(doctor_text)
    out["text_lower"] = out["text"].str.lower()

    return out, None

# ── FULL QUOTE HIGHLIGHT + RENDER ─────────────────────────────────────────────
def hl_text(text, focus=None):
    result = text
    order = (focus or []) + [t for t in THEMES if not focus or t not in focus]
    pairs = []
    for theme in order:
        css = HL_MAP.get(theme, "hl-def")
        for p in THEMES[theme]:
            pairs.append((p, css, len(p)))
    pairs.sort(key=lambda x: -x[2])
    for p, css, _ in pairs:
        result = re.sub(f'({re.escape(p)})', f'<span class="{css}">\\1</span>',
                        result, flags=re.IGNORECASE, count=3)
    return result

def quote_card(row, focus=None):
    is_c = str(row.get("setting","")).upper() in ["COMMUNITY","COMM"]
    css = "fq-card comm" if is_c else "fq-card acad"
    sbadge = f'<span class="seg-a">{row["setting"]}</span>' if is_c else f'<span class="seg-b">{row["setting"]}</span>'
    tags = "".join(f'<span class="tag" style="background:{TC.get(t,"#4a6080")}22;color:{TC.get(t,"#4a6080")};border:1px solid {TC.get(t,"#4a6080")}33">{t}</span>'
                   for t, pats in THEMES.items() if any(p in str(row.get("text_lower","")) for p in pats))
    h = hl_text(str(row.get("text","")), focus)
    st.markdown(f"""<div class="{css}">
        <div class="fq-meta">📎 <b style="color:#e2ecf8">{row["id"]}</b> &nbsp;{sbadge}&nbsp;
        <span>{row.get("specialty","")}</span> &nbsp; <span>{row.get("target","")}</span></div>
        <div class="fq-text">{h}</div>
        <div style="margin-top:10px;border-top:1px solid #1a2640;padding-top:6px">{tags}</div>
    </div>""", unsafe_allow_html=True)

# ── ANALYTICS ─────────────────────────────────────────────────────────────────
def mtch(tl, pats): return any(p in tl for p in pats)
def t_counts(df): return {t: int(df["text_lower"].apply(lambda x: mtch(x,p)).sum()) for t,p in THEMES.items()}
def d_counts(df): return {d: int(df["text_lower"].apply(lambda x: mtch(x,p)).sum()) for d,p in DRIVER_MAP.items()}

def bar_html(label, count, total, color="#3b6ef7"):
    pct = round(count/total*100) if total else 0
    st.markdown(f"""<div class="bar-row">
        <div class="bar-lbl">{label}</div>
        <div class="bar-trk"><div class="bar-fill" style="width:{pct}%;background:{color}"></div></div>
        <div class="bar-cnt">{count} <span style="color:#2a3a55">({pct}%)</span></div>
    </div>""", unsafe_allow_html=True)

# ── ENGINES ───────────────────────────────────────────────────────────────────
def co_occur(ta, tb, df):
    pa, pb = THEMES.get(ta,[ta.lower()]), THEMES.get(tb,[tb.lower()])
    T = len(df)
    both_rows = [row for _,row in df.iterrows() if any(p in row["text_lower"] for p in pa) and any(p in row["text_lower"] for p in pb)]
    nb = len(both_rows)
    oa = int(df["text_lower"].apply(lambda x: any(p in x for p in pa) and not any(p in x for p in pb)).sum())
    ob = int(df["text_lower"].apply(lambda x: any(p in x for p in pb) and not any(p in x for p in pa)).sum())
    return {"ta":ta,"tb":tb,"T":T,"oa":oa,"ob":ob,"nb":nb,"pct":round(nb/T*100) if T else 0,"rows":both_rows}

def cluster(anchor, df):
    pats = THEMES.get(anchor,[anchor.lower()])
    adf = df[df["text_lower"].apply(lambda x: any(p in x for p in pats))]
    n = len(adf)
    if n == 0: return {"anchor":anchor,"n":0,"T":len(df),"clusters":[],"avg":0,"verdict":"standalone","adf":adf}
    co = {}
    for t,tp in THEMES.items():
        if t==anchor: continue
        c = int(adf["text_lower"].apply(lambda x: any(p in x for p in tp)).sum())
        if c: co[t]=c
    avg = round(sum(len([t for t,tp in THEMES.items() if t!=anchor and any(p in row["text_lower"] for p in tp)]) for _,row in adf.iterrows())/n, 1)
    return {"anchor":anchor,"n":n,"T":len(df),"pct":round(n/len(df)*100),"clusters":sorted(co.items(),key=lambda x:-x[1]),"avg":avg,"verdict":"complex" if avg>=4 else "moderate" if avg>=2 else "standalone","adf":adf}

def complexity(anchor, df):
    pats = THEMES.get(anchor,[anchor.lower()])
    adf = df[df["text_lower"].apply(lambda x: any(p in x for p in pats))]
    n = len(adf)
    if n==0: return {"anchor":anchor,"n":0}
    solo,mod,comp=[],[],[]
    co=defaultdict(int)
    for _,row in adf.iterrows():
        others=[t for t,tp in THEMES.items() if t!=anchor and any(p in row["text_lower"] for p in tp)]
        for t in others: co[t]+=1
        rec=dict(row); rec["others"]=others; rec["n_other"]=len(others)
        if len(others)<=1: solo.append(rec)
        elif len(others)<=3: mod.append(rec)
        else: comp.append(rec)
    v="standalone" if len(solo)>len(comp)+len(mod) else "complex" if len(comp)>len(solo) else "mixed"
    return {"anchor":anchor,"n":n,"T":len(df),"solo":solo,"mod":mod,"comp":comp,
            "ns":len(solo),"nm":len(mod),"nc":len(comp),
            "pcts":round(len(solo)/n*100),"pctc":round(len(comp)/n*100),
            "top_co":dict(sorted(co.items(),key=lambda x:-x[1])[:8]),"verdict":v}

# ── INTENT / TOPICS ───────────────────────────────────────────────────────────
def topics(q):
    ql=q.lower()
    return list(dict.fromkeys([t for t,pats in THEMES.items() if any(p in ql for p in pats) or t.lower() in ql]))

def intent(q):
    ql=q.lower()
    tpcs=topics(q)
    if len(tpcs)>=2 and any(w in ql for w in ["hand in hand","together","alongside","linked","co-occur","also mention","both","went together"]): return "co_occur"
    if any(w in ql for w in ["what else","other drivers","other themes","tagged to","cluster","travel with","come with","associated with","what other","drivers tagged"]): return "cluster"
    if any(w in ql for w in ["standalone","straightforward","straight forward","only driver","simple driver","complex","entangled","on its own","by itself","always with","was it just","was it only","was pfs"]): return "complexity"
    rules = {
        "comparison":[r'(community|academic|neuro|medonc|heme|on tl|off tl|co-loc|segment|setting|specialty).*(vs|versus|compare|differ|more|less|between|than)|(vs|versus|compare|differ|difference|different).*(community|academic|setting|specialty|tier|segment|group)'],
        "frequency": [r'how many|how often|what (percentage|proportion|%)|how frequent|number of|count'],
        "driver":    [r'(top|main|key|primary|most common|biggest|major|most important).*(reason|driver|factor|barrier|concern)|(most tied|most associated|which driver|what drives|why do)'],
        "quotes":    [r'(quote|verbatim|exact words|what.*(say|said)|show me.*(quote|example)|give me.*(quote|full response|full quote|full transcript))'],
        "cost":      [r'(cost|afford|insurance|reimburs|copay|prior auth|formulary|coverage|financial|patient assist)'],
        "seizure":   [r'seizure'],
        "endpoint":  [r'(pfs|os|overall survival|progression free).*(prefer|meaningful|important|endpoint)|(pfs.*(vs|or|versus).*os)'],
        "radiation": [r'(delay|defer|avoid).*(radiation|chemo|radiotherapy|rt)'],
        "patient":   [r'(patient|eligib).*(character|profile|who|age|grade|mutation|resection)|ideal candidate|appropriate patient'],
        "barrier":   [r'(barrier|obstacle|challenge|concern|hesitat|reluctan).*(prescrib|adopt|use)|what stops|what prevents'],
    }
    sc=defaultdict(int)
    for i,pats in rules.items():
        for p in pats:
            if re.search(p,ql): sc[i]+=1
    return max(sc,key=sc.get) if sc else "content"

def is_comp(q):
    ql=q.lower()
    sw=["community","academic","neuro","medonc","med onc","hemeonc","heme","on tl","off tl","co-loc","tier","specialty","setting","segment","group"]
    cw=["vs","versus","compare","difference","different","more","less","between","than","how does","what changed","higher","lower"]
    return any(w in ql for w in sw) and any(w in ql for w in cw)

def detect_segs(q, df):
    ql=q.lower()
    for col in ["setting","target","specialty"]:
        vals=df[col].unique().tolist()
        hits=[v for v in vals if v.lower() in ql or v.lower().replace("-","").replace(" ","") in ql.replace("-","").replace(" ","")]
        if len(hits)>=2: return (col,hits[0]),(col,hits[1])
        if len(hits)==1:
            others=[v for v in vals if v!=hits[0]]
            if others: return (col,hits[0]),(col,others[0])
    s=df["setting"].unique().tolist()
    if len(s)>=2: return ("setting",s[0]),("setting",s[1])
    return None,None

# ── ANSWER ────────────────────────────────────────────────────────────────────
def answer(q, adf, fdf):
    T=len(fdf); itn=intent(q); tpcs=topics(q); comp=is_comp(q)
    qpats=[]
    for t in tpcs: qpats+=THEMES.get(t,[])
    if not qpats: qpats=[w for w in q.lower().split() if len(w)>4]
    res={"intent":itn,"topics":tpcs,"is_comp":comp,"n":len(adf),"T":T,"summary":"","chart":{},"rows":[],"comp":None,"co":None,"cl":None,"cx":None,"export":[]}

    if comp:
        sa,sb=detect_segs(q,fdf)
        if sa and sb:
            ca,va=sa; cb,vb=sb
            dfa=fdf[fdf[ca]==va]; dfb=fdf[fdf[cb]==vb]
            na,nb=len(dfa),len(dfb)
            tca,tcb=t_counts(dfa),t_counts(dfb)
            rows=[]
            for t in THEMES:
                a,b=tca.get(t,0),tcb.get(t,0)
                pa=round(a/na*100) if na else 0; pb=round(b/nb*100) if nb else 0
                rows.append({"Theme":t,f"{va}(n={na})":f"{a}({pa}%)",f"{vb}(n={nb})":f"{b}({pb}%)","D":f"+{pa-pb}%" if pa>pb else f"{pa-pb}%","_d":pa-pb,"_pa":pa,"_pb":pb,"_a":a,"_b":b})
            rows=([r for r in rows if r["Theme"] in tpcs] or rows) if tpcs else sorted(rows,key=lambda x:abs(x["_d"]),reverse=True)[:10]
            top=rows[0] if rows else {}
            if top:
                w=va if top["_d"]>0 else vb; l=vb if top["_d"]>0 else va
                res["summary"]=f"**{w}** mentions **{top['Theme']}** more than **{l}** by **{abs(top['_d'])}pp** ({top['_pa']}% vs {top['_pb']}%)."
            focus=qpats or THEMES.get(tpcs[0],[]) if tpcs else []
            res["comp"]={"rows":rows,"va":va,"vb":vb,"na":na,"nb":nb,"dfa":dfa,"dfb":dfb,"focus":focus}
            res["intent"]="comparison"; return res

    if itn=="co_occur" and len(tpcs)>=2:
        co=co_occur(tpcs[0],tpcs[1],fdf)
        res["summary"]=f"**{co['nb']} of {co['T']} respondents ({co['pct']}%)** mentioned both **{tpcs[0]}** and **{tpcs[1]}**."
        res["co"]=co; return res

    if itn=="cluster" and tpcs:
        cl=cluster(tpcs[0],fdf)
        res["summary"]=f"**{cl['n']} of {cl['T']} respondents ({cl.get('pct',0)}%)** mentioned **{tpcs[0]}**. Avg **{cl['avg']} other themes**. Complexity: **{cl['verdict']}**."
        res["cl"]=cl; return res

    if itn=="complexity" and tpcs:
        cx=complexity(tpcs[0],fdf)
        vl={"standalone":"🟢 Standalone","complex":"🔴 Complex / Entangled","mixed":"🟡 Mixed"}.get(cx.get("verdict","mixed"),"")
        res["summary"]=f"**{cx.get('n',0)} respondents** on **{tpcs[0]}**: **{cx.get('ns',0)} standalone** · **{cx.get('nc',0)} complex**. {vl}"
        res["cx"]=cx; return res

    mdf=adf[adf["text_lower"].apply(lambda x: any(p in x for p in qpats))] if qpats else adf
    nm=len(mdf); pct=round(nm/T*100) if T else 0; tstr=" + ".join(tpcs) if tpcs else "this topic"
    res["n"]=nm; res["rows"]=mdf.to_dict("records")
    res["export"]=[{"ID":r["id"],"Setting":r["setting"],"Target":r["target"],"Specialty":r["specialty"],"Full Response":r["text"]} for r in res["rows"]]

    if itn=="frequency":
        res["summary"]=f"**{nm} of {T} respondents ({pct}%)** mentioned {tstr}."
        res["chart"]={"Associated Themes":dict(sorted({t:v for t,v in t_counts(mdf).items() if v>0}.items(),key=lambda x:-x[1])[:10])}
    elif itn=="driver":
        dc=dict(sorted(d_counts(mdf).items(),key=lambda x:-x[1]))
        top=next(iter(dc.items()),("unknown",0))
        res["summary"]=f"Top driver for **{tstr}**: **{top[0]}** — cited by **{top[1]} respondents**."
        res["chart"]={"Driver Ranking":dc}
    elif itn=="cost":
        cdf=fdf[fdf["text_lower"].apply(lambda x: mtch(x,THEMES["Cost/Insurance"]))]
        cn=len(cdf)
        appr={"Prior Auth/Appeal":int(cdf["text_lower"].apply(lambda x: mtch(x,["prior auth","appeal","peer to peer"])).sum()),"Patient Assistance":int(cdf["text_lower"].apply(lambda x: mtch(x,["patient assistance","patient support","patient program"])).sum()),"Manufacturer Programs":int(cdf["text_lower"].apply(lambda x: mtch(x,["manufacturer","company","free drug"])).sum()),"Copay Support":int(cdf["text_lower"].apply(lambda x: mtch(x,["copay","co-pay"])).sum()),"Office/Pharmacist":int(cdf["text_lower"].apply(lambda x: mtch(x,["my office","pharmacist","staff"])).sum())}
        res["summary"]=f"**{cn} of {T} respondents ({round(cn/T*100) if T else 0}%)** discussed cost/insurance."
        res["chart"]={"Reimbursement Approaches":dict(sorted(appr.items(),key=lambda x:-x[1]))}
        res["rows"]=cdf.to_dict("records"); res["export"]=[{"ID":r["id"],"Setting":r["setting"],"Target":r.get("target",""),"Specialty":r["specialty"],"Full Response":r["text"]} for r in res["rows"]]
    elif itn=="seizure":
        sdf=fdf[fdf["text_lower"].apply(lambda x: mtch(x,THEMES["Seizures"]))]
        sn=len(sdf); pos=int(sdf["text_lower"].apply(lambda x: any(p in x for p in ["reduc","decreas","lower","control","fewer","less"])).sum()); neg=int(sdf["text_lower"].apply(lambda x: any(p in x for p in ["trigger","worsen","increase","risk","not primary"])).sum())
        res["summary"]=f"**{sn} of {T} respondents** discussed seizures. **{pos} benefit** · **{neg} concern**."
        res["chart"]={"Seizure Sentiment":{"Benefit":pos,"Concern":neg,"Neutral":max(0,sn-pos-neg)}}
        res["rows"]=sdf.to_dict("records"); res["export"]=[{"ID":r["id"],"Setting":r["setting"],"Target":r.get("target",""),"Specialty":r["specialty"],"Full Response":r["text"]} for r in res["rows"]]
    elif itn=="endpoint":
        pn=int(fdf["text_lower"].apply(lambda x: mtch(x,THEMES["PFS"])).sum()); on=int(fdf["text_lower"].apply(lambda x: mtch(x,THEMES["OS"])).sum()); bn=int(fdf["text_lower"].apply(lambda x: mtch(x,THEMES["PFS"]) and mtch(x,THEMES["OS"])).sum())
        res["summary"]=f"**PFS**: {pn} ({round(pn/T*100) if T else 0}%) · **OS**: {on} ({round(on/T*100) if T else 0}%) · **Both**: {bn}"
        res["chart"]={"Endpoint Split":{"PFS":pn,"OS":on,"Both":bn}}
        res["rows"]=fdf[fdf["text_lower"].apply(lambda x: mtch(x,THEMES["PFS"]))].to_dict("records")
        res["export"]=[{"ID":r["id"],"Setting":r["setting"],"Target":r.get("target",""),"Specialty":r["specialty"],"Full Response":r["text"]} for r in res["rows"]]
    elif itn=="radiation":
        rdf=fdf[fdf["text_lower"].apply(lambda x: mtch(x,THEMES["Radiation Delay"]))]; rn=len(rdf)
        why={"Cognitive Preservation":int(rdf["text_lower"].apply(lambda x: any(p in x for p in ["cognitive","neurocognit","brain fog","dementia"])).sum()),"Long-term Side Effects":int(rdf["text_lower"].apply(lambda x: "long term" in x or "long-term" in x).sum()),"Fertility":int(rdf["text_lower"].apply(lambda x: mtch(x,THEMES["Fertility"])).sum()),"Quality of Life":int(rdf["text_lower"].apply(lambda x: mtch(x,THEMES["Quality of Life"])).sum()),"Younger Patients":int(rdf["text_lower"].apply(lambda x: any(p in x for p in ["young","younger","less than 40"])).sum())}
        res["summary"]=f"**{rn} of {T} respondents** cited delaying radiation/chemo as a benefit."
        res["chart"]={"Why Delay Radiation":dict(sorted(why.items(),key=lambda x:-x[1]))}
        res["rows"]=rdf.to_dict("records"); res["export"]=[{"ID":r["id"],"Setting":r["setting"],"Target":r.get("target",""),"Specialty":r["specialty"],"Full Response":r["text"]} for r in res["rows"]]
    elif itn=="patient":
        ch={"IDH1/IDH2 Mutant":int(fdf["text_lower"].apply(lambda x: mtch(x,THEMES["IDH Mutation"])).sum()),"Grade 2 Disease":int(fdf["text_lower"].apply(lambda x: "grade 2" in x or "grade two" in x).sum()),"Residual/Recurring":int(fdf["text_lower"].apply(lambda x: any(p in x for p in ["residual tumor","residual disease","recurring"])).sum()),"Good Performance Status":int(fdf["text_lower"].apply(lambda x: any(p in x for p in ["performance status","ecog","kps"])).sum()),"Younger Patients":int(fdf["text_lower"].apply(lambda x: any(p in x for p in ["young","younger","less than 40"])).sum())}
        res["summary"]=f"Patient characteristics across {T} respondents:"
        res["chart"]={"Ideal Patient Profile":dict(sorted(ch.items(),key=lambda x:-x[1]))}
    elif itn=="barrier":
        bp=["concern","hesitant","not comfortable","challenge","issue","barrier","problem","limited","refused","denied"]
        bdf=mdf[mdf["text_lower"].apply(lambda x: any(p in x for p in bp))] if nm>0 else mdf
        res["summary"]=f"**{len(bdf)} respondents** expressed barriers around {tstr}."
        res["chart"]={"Barrier Drivers":dict(sorted(d_counts(bdf).items(),key=lambda x:-x[1])[:8])}
        res["rows"]=bdf.to_dict("records"); res["export"]=[{"ID":r["id"],"Setting":r["setting"],"Target":r.get("target",""),"Specialty":r["specialty"],"Full Response":r["text"]} for r in res["rows"]]
    else:
        res["summary"]=f"**{nm} of {T} respondents ({pct}%)** discussed {tstr}."
        res["chart"]={"Themes Found":dict(sorted({t:v for t,v in t_counts(mdf).items() if v>0}.items(),key=lambda x:-x[1])[:10])}
    return res


# ── DASHBOARD ENGINE ──────────────────────────────────────────────────────────

def build_segment_breakdown(theme, full_df):
    pats = THEMES.get(theme, [theme.lower()])
    result = {}
    for col in ["setting", "specialty", "target"]:
        breakdown = {}
        for val in full_df[col].dropna().unique():
            seg = full_df[full_df[col] == val]
            n_seg = len(seg)
            n_match = int(seg["text_lower"].apply(lambda x: any(p in x for p in pats)).sum())
            if n_seg > 0:
                breakdown[val] = {"n": n_match, "total": n_seg, "pct": round(n_match / n_seg * 100)}
        result[col] = breakdown
    return result


def build_exec_summary(theme, theme_df, full_df, seg_data):
    T = len(full_df); n = len(theme_df); pct = round(n/T*100) if T else 0
    points = []
    # 1 — Prevalence
    freq = "one of the most prominent" if pct >= 50 else "a notable" if pct >= 30 else "a less frequently cited"
    points.append(f"**{n} of {T} respondents ({pct}%)** mentioned **{theme}**, making it {freq} topic in the uploaded dataset.")
    # 2 — Setting split
    s = seg_data.get("setting", {})
    if len(s) >= 2:
        items = sorted(s.items(), key=lambda x: -x[1]["pct"])
        top, bot = items[0], items[-1]
        gap = top[1]["pct"] - bot[1]["pct"]
        if gap >= 10:
            points.append(f"**{top[0]}** HCPs discuss {theme} more than **{bot[0]}** ({top[1]['pct']}% vs {bot[1]['pct']}%) — a **{gap}pp gap** suggesting practice setting shapes how this topic surfaces.")
        else:
            points.append(f"Mentions are broadly consistent across settings — **{top[0]}**: {top[1]['pct']}% vs **{bot[0]}**: {bot[1]['pct']}% — no major setting-driven divide.")
    # 3 — Specialty split
    sp = seg_data.get("specialty", {})
    if len(sp) >= 2:
        items = sorted(sp.items(), key=lambda x: -x[1]["pct"])
        top_sp = items[0]
        others_str = "; ".join(f"{v}: {d['pct']}%" for v,d in items[1:])
        points.append(f"By specialty, **{top_sp[0]}** most frequently discusses {theme} ({top_sp[1]['pct']}% of that group). Other specialties: {others_str}.")
    # 4 — Co-theme
    co = {}
    for t, p2 in THEMES.items():
        if t == theme: continue
        c = int(theme_df["text_lower"].apply(lambda x: any(p in x for p in p2)).sum())
        if c: co[t] = c
    if co:
        top3 = sorted(co.items(), key=lambda x: -x[1])[:3]
        co_str = ", ".join(f"**{t}** ({c} respondents)" for t,c in top3)
        points.append(f"{theme} most frequently co-occurs with {co_str} — suggesting it is rarely discussed in isolation and is part of a broader clinical narrative.")
    # 5 — Treatment impact / complexity
    solo = sum(1 for _,row in theme_df.iterrows() if sum(1 for t,p2 in THEMES.items() if t!=theme and any(p in row["text_lower"] for p in p2)) <= 1)
    comp = sum(1 for _,row in theme_df.iterrows() if sum(1 for t,p2 in THEMES.items() if t!=theme and any(p in row["text_lower"] for p in p2)) >= 4)
    if solo > comp:
        points.append(f"**{theme} acts as a standalone treatment driver** — {solo} of {n} respondents ({round(solo/n*100)}%) cited it without many other themes, indicating it directly and independently influences prescribing decisions.")
    elif comp > solo:
        points.append(f"**{theme} is a complex, multi-factor driver** — {comp} of {n} respondents ({round(comp/n*100)}%) discussed it alongside 4+ other themes, suggesting it is embedded in a broader decision-making process rather than a single decisive factor.")
    else:
        points.append(f"**{theme} shows mixed complexity** — cited as a direct standalone reason by {solo} respondents and as part of multi-theme discussions by {comp} others, indicating context-dependent importance.")
    return points



# ─────────────────────────────────────────────────────────────────────────────
# SENTIMENT ENGINE — context-aware, pharma-tuned, zero hallucination
# Every classification traced to exact sentence + trigger words
# ─────────────────────────────────────────────────────────────────────────────

# Positive signals — HCP pharma language
SENT_POSITIVE = [
    "good efficacy","great efficacy","excellent efficacy","strong efficacy",
    "impressive data","impressive results","impressive benefit",
    "well tolerated","good tolerability","favorable profile","favorable side effect",
    "manageable","no significant side","minimal side","low toxicity","low side",
    "confident","comfortable prescribing","happy to prescribe","prefer",
    "proven benefit","proven efficacy","clear benefit","significant benefit",
    "significant improvement","meaningful improvement",
    "easy to use","easy to administer","convenient","simple",
    "good quality of life","improved quality","maintain quality",
    "delay radiation","delay chemo","avoid radiation","avoid chemo","defer radiation",
    "targeted therapy","precise","specific","rational",
    "no issues","no problem","no concern","no barrier","no challenge",
    "no reimbursement issue","covered","approved","on formulary",
    "patients do well","patients respond","patients benefit",
    "supports","support","endorse","recommend","choose","select",
    "first choice","go to","standard","routine",
    "promising","encouraging","positive data","positive results",
    "effective","efficacious","works well","good outcomes",
]

# Negative signals — HCP pharma language
SENT_NEGATIVE = [
    "concern","concerned about","worry","worried","hesitant","hesitation",
    "reluctant","reluctance","not comfortable","uncomfortable",
    "barrier","obstacle","challenge","difficult","difficulty",
    "issue with","problem with","limitation","limited by",
    "uncertain","uncertainty","unclear","not sure","unsure","not confident",
    "adverse event","toxicity concern","side effect concern","toxicity issue",
    "not tolerated","poorly tolerated","intolerable","significant toxicity",
    "insurance denial","denied","prior auth","reimbursement challenge",
    "access issue","affordability","can't afford","too expensive","high cost",
    "not covered","formulary issue","off formulary",
    "don't prescribe","do not prescribe","avoid prescribing","rarely prescribe",
    "not appropriate","inappropriate","not indicated","not eligible",
    "liver dysfunction","hepatotoxicity","liver concern",
    "fertility concern","birth defect","pregnancy concern",
    "not durable","short duration","limited duration",
    "os data missing","no os data","os not available","overall survival not",
    "still early","not enough data","insufficient data","limited data",
    "patient refuses","patient reluctant","patient concern",
    "not first choice","second choice","last resort",
]

# Negation prefixes — flip negative to positive when present before trigger
NEGATION_PREFIXES = [
    "no ", "not ", "never ", "without ",
    "don't have ", "do not have ", "doesn't have ", "does not have ",
    "haven't had any ", "haven't had ", "have not had ", "hadn't had ",
    "haven't seen ", "have not seen ", "haven't experienced ", "have not experienced ",
    "i haven't ", "i have not ", "we haven't ", "we have not ",
    "hasn't been ", "has not been ", "wasn't ", "was not ",
    "no significant ", "no major ", "no real ", "no actual ",
    "eliminates ", "removes ", "resolves ", "no longer ",
    "absence of ", "free of ", "free from ",
]

# Context words that make negative words POSITIVE (e.g. "delay radiation" = good)
POSITIVE_CONTEXT_OVERRIDE = {
    "delay": ["radiation","chemo","chemotherapy","radiotherapy","alkylat","rt","surgery"],
    "avoid": ["radiation","chemo","chemotherapy","toxicity","side effect","hospitalization"],
    "prevent": ["progression","recurrence","relapse","tumor growth","seizure"],
    "reduce": ["seizure","tumor","progression","side effect","toxicity","risk"],
    "eliminate": ["disease","tumor","progression"],
    "control": ["seizure","tumor","disease","symptom"],
    "protect": ["cognition","fertility","quality","brain"],
}


def classify_sentence_sentiment(sentence):
    """
    Classify a single sentence.
    Returns: (sentiment, pos_triggers, neg_triggers, confidence)
    sentiment: POSITIVE / NEGATIVE / MIXED / NEUTRAL
    confidence: HIGH / MEDIUM / LOW
    """
    sl = sentence.lower()

    pos_hits = []
    neg_hits = []

    # ── Check positive signals ────────────────────────────────────────────
    for sig in SENT_POSITIVE:
        if sig in sl:
            pos_hits.append(sig)

    # ── Check negative signals with negation detection ────────────────────
    for sig in SENT_NEGATIVE:
        if sig in sl:
            # Check if a negation prefix appears before this signal
            sig_idx = sl.find(sig)
            prefix_window = sl[max(0, sig_idx-50):sig_idx]
            negated = any(neg in prefix_window for neg in NEGATION_PREFIXES)
            if negated:
                # Negated negative = positive signal
                pos_hits.append(f"no {sig}")
            else:
                neg_hits.append(sig)

    # ── Context override — certain words are positive in pharma context ────
    for word, contexts in POSITIVE_CONTEXT_OVERRIDE.items():
        if word in sl:
            if any(ctx in sl for ctx in contexts):
                # This is a positive usage (e.g. "delay radiation")
                # Remove from neg_hits if it was added
                neg_hits = [h for h in neg_hits if word not in h]
                if word + " (positive context)" not in pos_hits:
                    pos_hits.append(word + " (positive context)")

    # Deduplicate
    pos_hits = list(dict.fromkeys(pos_hits))[:3]
    neg_hits = list(dict.fromkeys(neg_hits))[:3]

    # ── Classify ──────────────────────────────────────────────────────────
    if pos_hits and neg_hits:
        sentiment = "MIXED"
        confidence = "HIGH" if (len(pos_hits) >= 2 or len(neg_hits) >= 2) else "MEDIUM"
    elif pos_hits:
        sentiment = "POSITIVE"
        confidence = "HIGH" if len(pos_hits) >= 2 else "MEDIUM"
    elif neg_hits:
        sentiment = "NEGATIVE"
        confidence = "HIGH" if len(neg_hits) >= 2 else "MEDIUM"
    else:
        sentiment = "NEUTRAL"
        confidence = "LOW"

    return sentiment, pos_hits, neg_hits, confidence


def run_sentiment_analysis(theme, full_df, theme_filter=None):
    """
    Run sentence-level sentiment analysis on SPEAKER_B responses
    for respondents who mentioned the given theme.
    Returns structured results with full traceability.
    """
    pats = THEMES.get(theme, [theme.lower()]) if theme != "ALL" else []

    # Filter to relevant respondents
    if pats:
        rel_df = full_df[full_df["text_lower"].apply(lambda x: any(p in x for p in pats))]
    else:
        rel_df = full_df

    results = []  # one entry per classified sentence

    for _, row in rel_df.iterrows():
        text = str(row.get("text", ""))
        # Split into sentences
        sents = [s.strip() for s in re.split(r'(?<=[.!?])\s+', text) if len(s.strip()) > 20]

        for sent in sents:
            # Only classify sentences that reference the theme (or all if no theme)
            sl = sent.lower()
            theme_relevant = True
            if pats:
                theme_relevant = any(p in sl for p in pats)

            sentiment, pos_hits, neg_hits, confidence = classify_sentence_sentiment(sent)

            # Detect which themes appear in this sentence
            sent_themes = [t for t, tp in THEMES.items() if any(p in sl for p in tp)]

            results.append({
                "id": row["id"],
                "setting": row["setting"],
                "specialty": row["specialty"],
                "target": row.get("target", ""),
                "sentence": sent,
                "sentiment": sentiment,
                "pos_triggers": pos_hits,
                "neg_triggers": neg_hits,
                "confidence": confidence,
                "theme_relevant": theme_relevant,
                "sentence_themes": sent_themes,
            })

    # Aggregate counts
    theme_sents = [r for r in results if r["theme_relevant"]]
    total_sents = len(theme_sents)

    counts = {"POSITIVE": 0, "NEGATIVE": 0, "MIXED": 0, "NEUTRAL": 0}
    for r in theme_sents:
        counts[r["sentiment"]] += 1

    # Per-segment breakdown
    seg_sentiment = {}
    for col in ["setting", "specialty", "target"]:
        seg_sentiment[col] = {}
        vals = set(r[col] for r in theme_sents if r[col])
        for val in vals:
            val_sents = [r for r in theme_sents if r[col] == val]
            vt = len(val_sents)
            # Count unique respondents (HCP IDs) in this segment
            n_resp = len(set(r["id"] for r in val_sents))
            seg_sentiment[col][val] = {
                "total": vt,           # total sentences
                "n_resp": n_resp,       # unique respondents
                "pos": sum(1 for r in val_sents if r["sentiment"] == "POSITIVE"),
                "neg": sum(1 for r in val_sents if r["sentiment"] == "NEGATIVE"),
                "mixed": sum(1 for r in val_sents if r["sentiment"] == "MIXED"),
                "neutral": sum(1 for r in val_sents if r["sentiment"] == "NEUTRAL"),
            }

    # Top positive and negative sentences
    pos_examples = sorted([r for r in theme_sents if r["sentiment"] == "POSITIVE"],
                          key=lambda x: len(x["pos_triggers"]), reverse=True)[:5]
    neg_examples = sorted([r for r in theme_sents if r["sentiment"] == "NEGATIVE"],
                          key=lambda x: len(x["neg_triggers"]), reverse=True)[:5]
    mixed_examples = [r for r in theme_sents if r["sentiment"] == "MIXED"][:4]

    # Most common triggers
    all_pos_triggers = []
    all_neg_triggers = []
    for r in theme_sents:
        all_pos_triggers.extend(r["pos_triggers"])
        all_neg_triggers.extend(r["neg_triggers"])

    from collections import Counter
    top_pos_triggers = dict(Counter(all_pos_triggers).most_common(8))
    top_neg_triggers = dict(Counter(all_neg_triggers).most_common(8))

    return {
        "theme": theme,
        "n_respondents": len(rel_df),
        "total_sentences": total_sents,
        "counts": counts,
        "seg_sentiment": seg_sentiment,
        "pos_examples": pos_examples,
        "neg_examples": neg_examples,
        "mixed_examples": mixed_examples,
        "top_pos_triggers": top_pos_triggers,
        "top_neg_triggers": top_neg_triggers,
        "all_results": theme_sents,
    }


def render_sentiment_tab(theme, full_df):
    """Render the sentiment analysis tab inside the dashboard."""
    with st.spinner("Running sentence-level sentiment analysis..."):
        sa = run_sentiment_analysis(theme, full_df)

    T_sents = sa["total_sentences"]
    if T_sents == 0:
        st.warning(f"No sentences referencing {theme} found.")
        return

    counts = sa["counts"]

    # ── Overall sentiment split ───────────────────────────────────────────
    st.markdown('<div class="sec-lbl">OVERALL SENTIMENT SPLIT — SENTENCE LEVEL</div>', unsafe_allow_html=True)
    st.markdown(f'<div style="font-size:11px;color:#4a6080;margin-bottom:14px">Based on {T_sents} sentences referencing <b style="color:#e2ecf8">{theme}</b> across {sa["n_respondents"]} respondents. Each sentence classified independently. Every classification backed by trigger words from the text.</div>', unsafe_allow_html=True)

    # Stat cards
    c1,c2,c3,c4 = st.columns(4)
    card_configs = [
        ("✅ POSITIVE", counts["POSITIVE"], "#10b981", "Favourable language about this theme"),
        ("❌ NEGATIVE", counts["NEGATIVE"], "#ef4444", "Concern, barrier, hesitation language"),
        ("🔀 MIXED", counts["MIXED"], "#f59e0b", "Both positive & negative in same sentence"),
        ("⬜ NEUTRAL", counts["NEUTRAL"], "#64748b", "Clinical/factual — no evaluative language"),
    ]
    for col, (label, count, color, desc) in zip([c1,c2,c3,c4], card_configs):
        pct = round(count/T_sents*100) if T_sents else 0
        col.markdown(f'''<div class="card" style="text-align:center;border-left:3px solid {color}">
            <div style="font-size:11px;color:{color};font-weight:600;margin-bottom:4px">{label}</div>
            <div class="stat-num" style="color:{color};font-size:32px">{count}</div>
            <div style="font-family:'IBM Plex Mono',monospace;font-size:13px;color:#e2ecf8;margin-top:2px">{pct}%</div>
            <div style="font-size:9px;color:#4a6080;margin-top:4px">{desc}</div>
        </div>''', unsafe_allow_html=True)

    # ── Net Sentiment Score ──────────────────────────────────────────────
    pos_n  = counts["POSITIVE"]
    neg_n  = counts["NEGATIVE"]
    mix_n  = counts["MIXED"]
    neu_n  = counts["NEUTRAL"]
    total_mentions = pos_n + neg_n + neu_n  # denominator per formula
    nss = round((pos_n - neg_n) / total_mentions * 100, 1) if total_mentions > 0 else 0
    nss_color  = "#10b981" if nss > 10 else "#ef4444" if nss < -10 else "#f59e0b"
    nss_label  = "Net positive" if nss > 10 else "Net negative" if nss < -10 else "Balanced"
    nss_interp = (
        f"More positive than negative sentiment on {theme}. Doctors broadly favour this topic."
        if nss > 10 else
        f"More negative than positive sentiment on {theme}. Concerns outweigh favourable mentions."
        if nss < -10 else
        f"Balanced sentiment on {theme}. Positive and negative mentions are roughly equal."
    )
    st.markdown(f'''<div class="card" style="border-left:4px solid {nss_color};margin-top:4px">
        <div style="display:flex;align-items:center;gap:20px;flex-wrap:wrap">
            <div style="text-align:center;min-width:100px">
                <div style="font-size:10px;color:{nss_color};font-weight:600;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:4px">Net Sentiment Score</div>
                <div style="font-family:'IBM Plex Mono',monospace;font-size:42px;font-weight:700;color:{nss_color};line-height:1">{nss:+.1f}</div>
                <div style="font-size:11px;color:{nss_color};margin-top:2px">{nss_label}</div>
            </div>
            <div style="flex:1;min-width:200px">
                <div style="font-size:12px;color:#e2ecf8;line-height:1.6;margin-bottom:8px">{nss_interp}</div>
                <div style="font-size:11px;color:#4a6080;line-height:1.6">
                    <b style="color:#8a9ab5">Formula:</b> (Positive − Negative) ÷ Total Mentions × 100<br>
                    <b style="color:#8a9ab5">Where:</b> Total Mentions = Positive ({pos_n}) + Negative ({neg_n}) + Neutral ({neu_n})<br>
                    <b style="color:#8a9ab5">Excludes:</b> Mixed sentences from denominator to avoid double-counting
                </div>
            </div>
            <div style="text-align:center;min-width:140px">
                <div style="font-size:10px;color:#4a6080;margin-bottom:6px">Score interpretation</div>
                <div style="font-size:11px;line-height:1.9">
                    <span style="color:#10b981">▲ &gt; +10</span> — Net positive<br>
                    <span style="color:#f59e0b">● −10 to +10</span> — Balanced<br>
                    <span style="color:#ef4444">▼ &lt; −10</span> — Net negative
                </div>
            </div>
        </div>
    </div>''', unsafe_allow_html=True)

    # Sentiment bar
    st.markdown("<br>", unsafe_allow_html=True)
    pos_pct = round(counts["POSITIVE"]/T_sents*100) if T_sents else 0
    neg_pct = round(counts["NEGATIVE"]/T_sents*100) if T_sents else 0
    mix_pct = round(counts["MIXED"]/T_sents*100) if T_sents else 0
    neu_pct = 100 - pos_pct - neg_pct - mix_pct

    st.markdown(f'''<div style="margin-bottom:20px">
        <div style="font-size:11px;color:#4a6080;margin-bottom:6px">Sentiment distribution across {T_sents} sentences</div>
        <div style="display:flex;height:20px;border-radius:4px;overflow:hidden;gap:2px">
            <div style="width:{pos_pct}%;background:#10b981;display:flex;align-items:center;justify-content:center">
                <span style="font-size:9px;color:white;font-weight:600">{pos_pct}%</span></div>
            <div style="width:{neg_pct}%;background:#ef4444;display:flex;align-items:center;justify-content:center">
                <span style="font-size:9px;color:white;font-weight:600">{neg_pct}%</span></div>
            <div style="width:{mix_pct}%;background:#f59e0b;display:flex;align-items:center;justify-content:center">
                <span style="font-size:9px;color:white;font-weight:600">{mix_pct}%</span></div>
            <div style="flex:1;background:#334155;display:flex;align-items:center;justify-content:center">
                <span style="font-size:9px;color:#94a3b8;font-weight:600">{neu_pct}%</span></div>
        </div>
        <div style="display:flex;gap:16px;margin-top:6px">
            <span style="font-size:10px;color:#10b981">■ Positive</span>
            <span style="font-size:10px;color:#ef4444">■ Negative</span>
            <span style="font-size:10px;color:#f59e0b">■ Mixed</span>
            <span style="font-size:10px;color:#64748b">■ Neutral</span>
        </div>
    </div>''', unsafe_allow_html=True)

    # ── Segment breakdown ─────────────────────────────────────────────────
    st.markdown('<div class="sec-lbl">SENTIMENT BY SEGMENT</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:14px">How sentiment on this theme breaks down across each segment. NSS = Net Sentiment Score — positive means more favourable than negative mentions.</div>', unsafe_allow_html=True)

    for seg_key, seg_label in [("setting","Practice Setting"),("specialty","Specialty"),("target","Target Type")]:
        seg_data = sa["seg_sentiment"].get(seg_key, {})
        if not seg_data: continue

        st.markdown(f'''<div class="card" style="margin-bottom:14px">
            <div class="sec-lbl" style="margin-bottom:14px">{seg_label}</div>
            <div style="display:grid;grid-template-columns:160px 1fr 70px 70px 70px 80px 80px 90px;gap:0;font-size:10px;font-weight:600;color:#4a6080;letter-spacing:0.5px;text-transform:uppercase;border-bottom:1px solid #1a2640;padding-bottom:6px;margin-bottom:6px">
                <span>Segment</span>
                <span>Sentiment bar</span>
                <span style="text-align:center">Positive</span>
                <span style="text-align:center">Negative</span>
                <span style="text-align:center">Mixed</span>
                <span style="text-align:center">HCPs</span>
                <span style="text-align:center">Sentences</span>
                <span style="text-align:center">NSS</span>
            </div>''', unsafe_allow_html=True)

        for val, d in sorted(seg_data.items(), key=lambda x: -(x[1]["pos"]+x[1]["neg"])):
            vt = d["total"]
            if vt == 0: continue
            pos_n = d["pos"]; neg_n = d["neg"]; mix_n = d.get("mixed",0)
            n_resp = d.get("n_resp", "—")
            p = round(pos_n/vt*100); n = round(neg_n/vt*100); m = round(mix_n/vt*100)
            seg_nss_denom = pos_n + neg_n + d.get("neutral", 0)
            seg_nss = round((pos_n - neg_n) / seg_nss_denom * 100) if seg_nss_denom > 0 else 0
            nss_c = "#10b981" if seg_nss > 10 else "#ef4444" if seg_nss < -10 else "#f59e0b"
            nss_label_s = "Net positive" if seg_nss > 10 else "Net negative" if seg_nss < -10 else "Balanced"
            pos_str = f"{p}%" if p > 0 else "—"
            neg_str = f"{n}%" if n > 0 else "—"
            mix_str = f"{m}%" if m > 0 else "—"

            st.markdown(f'''<div style="display:grid;grid-template-columns:160px 1fr 70px 70px 70px 80px 80px 90px;gap:0;align-items:center;padding:8px 0;border-bottom:1px solid #0f1823">
                <div style="font-size:12px;color:#e2ecf8;font-weight:500;padding-right:8px">{val[:22]}</div>
                <div style="display:flex;height:12px;border-radius:3px;overflow:hidden;gap:2px;margin-right:8px">
                    <div style="width:{p}%;background:#10b981;border-radius:2px 0 0 2px"></div>
                    <div style="width:{n}%;background:#ef4444"></div>
                    <div style="width:{m}%;background:#f59e0b"></div>
                    <div style="flex:1;background:#1e293b;border-radius:0 2px 2px 0"></div>
                </div>
                <div style="text-align:center;font-size:12px;color:#10b981;font-weight:500">{pos_str}</div>
                <div style="text-align:center;font-size:12px;color:{"#ef4444" if n > 0 else "#4a6080"};font-weight:500">{neg_str}</div>
                <div style="text-align:center;font-size:12px;color:{"#f59e0b" if m > 0 else "#4a6080"}">{mix_str}</div>
                <div style="text-align:center">
                    <div style="font-size:13px;color:#e2ecf8;font-weight:500">{n_resp}</div>
                    <div style="font-size:9px;color:#4a6080;margin-top:1px">HCPs</div>
                </div>
                <div style="text-align:center">
                    <div style="font-size:13px;color:#64748b">{vt}</div>
                    <div style="font-size:9px;color:#4a6080;margin-top:1px">sentences</div>
                </div>
                <div style="text-align:center">
                    <span style="font-family:'IBM Plex Mono',monospace;font-size:13px;font-weight:700;color:{nss_c}">{seg_nss:+d}</span>
                    <div style="font-size:9px;color:{nss_c};margin-top:1px">{nss_label_s}</div>
                </div>
            </div>''', unsafe_allow_html=True)

        # ── Calculation workings ──────────────────────────────────────────
        st.markdown('<div style="margin-top:14px;border-top:1px solid #1a2640;padding-top:12px">', unsafe_allow_html=True)
        st.markdown('<div style="font-size:10px;font-weight:600;color:#4a6080;letter-spacing:1px;text-transform:uppercase;margin-bottom:8px">NSS calculation workings</div>', unsafe_allow_html=True)

        # Header row
        st.markdown('''<div style="display:grid;grid-template-columns:160px 50px 50px 50px 50px 1fr 90px;gap:4px;font-size:9px;font-weight:600;color:#4a6080;letter-spacing:0.5px;text-transform:uppercase;padding:4px 8px;background:#0a1220;border-radius:6px 6px 0 0">
            <span>Segment</span>
            <span style="text-align:center">Pos</span>
            <span style="text-align:center">Neg</span>
            <span style="text-align:center">Neutral</span>
            <span style="text-align:center">Total</span>
            <span style="text-align:center">Formula: (Pos − Neg) ÷ Total × 100</span>
            <span style="text-align:center">NSS</span>
        </div>''', unsafe_allow_html=True)

        for val, d in sorted(seg_data.items(), key=lambda x: -(x[1]["pos"]+x[1]["neg"])):
            vt2 = d["total"]
            if vt2 == 0: continue
            pos2 = d["pos"]; neg2 = d["neg"]; neu2 = d.get("neutral", 0)
            denom2 = pos2 + neg2 + neu2
            nss2 = round((pos2 - neg2) / denom2 * 100) if denom2 > 0 else 0
            nss2_c = "#10b981" if nss2 > 10 else "#ef4444" if nss2 < -10 else "#f59e0b"
            formula_str = f"({pos2} − {neg2}) ÷ ({pos2} + {neg2} + {neu2}) × 100 = ({pos2 - neg2}) ÷ {denom2} × 100"
            # Sample size warning
            n_resp2 = d.get("n_resp", 0)
            warn = " ⚠" if (n_resp2 < 5 or vt2 < 10) else ""
            st.markdown(f'''<div style="display:grid;grid-template-columns:160px 50px 50px 50px 50px 1fr 90px;gap:4px;font-size:11px;align-items:center;padding:5px 8px;border-bottom:1px solid #0f1823;background:#0d1420">
                <span style="color:#c8d4e8;font-weight:500">{val[:22]}{warn}</span>
                <span style="text-align:center;color:#10b981;font-family:'IBM Plex Mono',monospace">{pos2}</span>
                <span style="text-align:center;color:{"#ef4444" if neg2 > 0 else "#4a6080"};font-family:'IBM Plex Mono',monospace">{neg2}</span>
                <span style="text-align:center;color:#4a6080;font-family:'IBM Plex Mono',monospace">{neu2}</span>
                <span style="text-align:center;color:#64748b;font-family:'IBM Plex Mono',monospace">{denom2}</span>
                <span style="color:#64748b;font-size:10px;font-family:'IBM Plex Mono',monospace;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{formula_str}</span>
                <span style="text-align:center;font-family:'IBM Plex Mono',monospace;font-weight:700;color:{nss2_c}">{nss2:+d}</span>
            </div>''', unsafe_allow_html=True)

        # Warning note if any small samples
        small_segs = [val for val, d in seg_data.items() if d.get("n_resp",0) < 5 or d.get("total",0) < 10]
        if small_segs:
            st.markdown(f'<div style="font-size:10px;color:#f59e0b;padding:6px 8px;background:#1a1200;border-radius:0 0 6px 6px">⚠ Small sample: {", ".join(small_segs[:3])} — treat NSS directionally, not conclusively (n &lt; 5 HCPs or &lt; 10 sentences)</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div style="height:4px;background:#0a1220;border-radius:0 0 6px 6px"></div>', unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # ── NSS methodology note ──────────────────────────────────────────────
    st.markdown(f'''<div class="card" style="border-left:3px solid #3b6ef7">
        <div class="sec-lbl">HOW THE NET SENTIMENT SCORE (NSS) WAS CALCULATED</div>
        <div style="font-size:13px;color:#c8d4e8;line-height:1.85">
            Each SPEAKER_B response was split into individual sentences. Only sentences referencing
            <b style="color:#e2ecf8">{theme}</b> were included. Each sentence was independently classified as
            <b style="color:#10b981">Positive</b>, <b style="color:#ef4444">Negative</b>,
            <b style="color:#f59e0b">Mixed</b>, or <b style="color:#64748b">Neutral</b>
            using a pharma-tuned lexicon of 60+ positive and 65+ negative signals,
            with negation detection (e.g. "no concern" → positive) and context overrides
            (e.g. "delay radiation" → positive).
        </div>
        <div style="margin-top:12px;background:#0a1220;border-radius:8px;padding:12px 16px">
            <div style="font-size:12px;color:#4a6080;margin-bottom:6px">Formula applied:</div>
            <div style="font-family:'IBM Plex Mono',monospace;font-size:13px;color:#e2ecf8">
                NSS = (Positive sentences − Negative sentences) ÷ Total sentences × 100
            </div>
            <div style="font-size:11px;color:#4a6080;margin-top:6px">
                Where Total = Positive + Negative + Neutral &nbsp;|&nbsp;
                Mixed sentences excluded from denominator to avoid double-counting &nbsp;|&nbsp;
                Score ranges from −100 (entirely negative) to +100 (entirely positive)
            </div>
        </div>
        <div style="margin-top:10px;display:flex;gap:20px;font-size:12px">
            <span><b style="color:#10b981">▲ &gt; +10</b> = Net positive sentiment</span>
            <span><b style="color:#f59e0b">● −10 to +10</b> = Balanced sentiment</span>
            <span><b style="color:#ef4444">▼ &lt; −10</b> = Net negative sentiment</span>
        </div>
    </div>''', unsafe_allow_html=True)

    # ── Verbatim evidence by sentiment ────────────────────────────────────
    for sent_type, examples, color, icon, label in [
        ("POSITIVE", sa["pos_examples"], "#10b981", "✅", "POSITIVE SENTENCES — what drives favourable sentiment"),
        ("NEGATIVE", sa["neg_examples"], "#ef4444", "❌", "NEGATIVE SENTENCES — what drives concern or hesitation"),
        ("MIXED", sa["mixed_examples"], "#f59e0b", "🔀", "MIXED SENTENCES — both favourable and concern language"),
    ]:
        if not examples:
            continue
        st.markdown(f'<div class="card"><div class="sec-lbl" style="color:{color}">{icon} {label}</div>', unsafe_allow_html=True)
        st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:12px">Full sentences from uploaded data. Trigger words shown below each quote.</div>', unsafe_allow_html=True)
        for ex in examples:
            is_comm = "COMMUNITY" in ex["setting"].upper() or "PRIVATE" in ex["setting"].upper()
            border_color = "#f59e0b" if is_comm else "#3b6ef7"
            seg_badge = f'<span class="seg-a">{ex["setting"][:20]}</span>' if is_comm else f'<span class="seg-b">{ex["setting"][:20]}</span>'
            # Highlight trigger words in sentence
            h_sent = ex["sentence"]
            for trg in ex["pos_triggers"] + ex["neg_triggers"]:
                clean_trg = trg.replace(" (positive context)", "")
                h_sent = re.sub(f'({re.escape(clean_trg)})',
                               f'<span style="background:{color}33;padding:1px 3px;border-radius:3px;color:{color};font-weight:600">\\1</span>',
                               h_sent, flags=re.IGNORECASE, count=1)
            triggers_html = ""
            if ex["pos_triggers"]:
                triggers_html += f'<span style="font-size:9px;color:#10b981">✅ {", ".join(ex["pos_triggers"][:2])}</span> &nbsp;' 
            if ex["neg_triggers"]:
                triggers_html += f'<span style="font-size:9px;color:#ef4444">❌ {", ".join(ex["neg_triggers"][:2])}</span>'
            st.markdown(f'''<div style="background:#0a1220;border-left:3px solid {color};border-radius:0 10px 10px 0;padding:12px 16px;margin-bottom:8px">
                <div style="font-family:'IBM Plex Mono',monospace;font-size:10px;color:#4a6080;margin-bottom:5px">
                    📎 {ex["id"]} &nbsp;{seg_badge}&nbsp;
                    <span style="color:#8a9ab5">{ex["specialty"][:25]}</span>
                </div>
                <div style="font-size:13px;color:#c8d4e8;font-style:italic;line-height:1.7">"{h_sent}"</div>
                <div style="margin-top:6px">{triggers_html}</div>
            </div>''', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Export ────────────────────────────────────────────────────────────
    st.markdown('<div class="sec-lbl">⬇ EXPORT SENTIMENT DATA</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:12px">Every sentence · HCP ID · Setting · Specialty · Target · Sentiment · Confidence · Trigger words · Net Sentiment Score per HCP</div>', unsafe_allow_html=True)

    # Build per-HCP NSS summary
    hcp_rows = {}
    for r in sa["all_results"]:
        hid = r["id"]
        if hid not in hcp_rows:
            hcp_rows[hid] = {"id":hid,"setting":r["setting"],"specialty":r["specialty"],
                             "target":r["target"],"pos":0,"neg":0,"mix":0,"neu":0,"total":0}
        hcp_rows[hid][r["sentiment"].lower()[:3]] = hcp_rows[hid].get(r["sentiment"].lower()[:3],0) + 1
        hcp_rows[hid]["total"] += 1

    def hcp_nss(h):
        denom = h["pos"] + h["neg"] + h["neu"]
        return round((h["pos"] - h["neg"]) / denom * 100, 1) if denom > 0 else 0

    hcp_nss_map = {h["id"]: hcp_nss(h) for h in hcp_rows.values()}

    # Sentence-level export
    export_rows = []
    for r in sa["all_results"]:
        hn = hcp_nss_map.get(r["id"], 0)
        export_rows.append({
            "HCP_ID": r["id"],
            "Practice_Setting": r["setting"],
            "Specialty": r["specialty"],
            "Target_Type": r["target"],
            "Sentence": r["sentence"],
            "Sentiment": r["sentiment"],
            "Confidence": r["confidence"],
            "Positive_Triggers": ", ".join(r["pos_triggers"]),
            "Negative_Triggers": ", ".join(r["neg_triggers"]),
            "Themes_in_Sentence": ", ".join(r["sentence_themes"][:5]),
            "HCP_NSS": f"{hn:+.1f}",
            "HCP_NSS_Label": "Net positive" if hn > 10 else "Net negative" if hn < -10 else "Balanced",
        })

    # HCP-level NSS summary
    hcp_summary_rows = []
    for h in hcp_rows.values():
        hn = hcp_nss(h)
        hcp_summary_rows.append({
            "HCP_ID": h["id"],
            "Practice_Setting": h["setting"],
            "Specialty": h["specialty"],
            "Target_Type": h["target"],
            "Total_Sentences": h["total"],
            "Positive": h["pos"],
            "Negative": h["neg"],
            "Mixed": h.get("mix",0),
            "Neutral": h["neu"],
            "Net_Sentiment_Score (NSS)": f"{hn:+.1f}",
            "NSS_Label": "Net positive" if hn > 10 else "Net negative" if hn < -10 else "Balanced",
        })

    if export_rows:
        tab_sent, tab_hcp = st.tabs(["📄 All sentences", "👤 HCP-level NSS summary"])

        with tab_sent:
            dfe = pd.DataFrame(export_rows)
            preview_cols = ["HCP_ID","Practice_Setting","Specialty","Sentiment","Confidence","HCP_NSS","Sentence"]
            st.dataframe(dfe[preview_cols].head(25), hide_index=True, use_container_width=True)
            st.download_button(
                f"⬇ Download all {len(export_rows)} sentences — {theme} sentiment CSV",
                dfe.to_csv(index=False).encode(),
                f"sentiment_sentences_{theme.replace('/','_')}.csv", "text/csv"
            )

        with tab_hcp:
            dfs = pd.DataFrame(hcp_summary_rows).sort_values("Net_Sentiment_Score (NSS)", ascending=False)
            st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:10px">One row per HCP. NSS = (Positive − Negative) ÷ (Pos + Neg + Neutral) × 100. Sorted most positive first.</div>', unsafe_allow_html=True)
            st.dataframe(dfs, hide_index=True, use_container_width=True)
            st.download_button(
                f"⬇ Download HCP-level NSS summary — {theme}",
                dfs.to_csv(index=False).encode(),
                f"sentiment_hcp_nss_{theme.replace('/','_')}.csv", "text/csv"
            )




def render_overall_sentiment(full_df):
    """
    Overall sentiment dashboard — all 18 themes, NSS by segment.
    No quotes shown. Stats only. Download as Excel.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    st.markdown('''<div style="background:#0d1420;border:1px solid #1a2640;border-radius:14px;padding:18px 22px;margin-bottom:20px;border-left:4px solid #8b5cf6">
        <div style="font-size:10px;letter-spacing:2px;text-transform:uppercase;color:#8b5cf6;font-weight:600;margin-bottom:6px;font-family:'IBM Plex Mono',monospace">SENTIMENT OVERVIEW</div>
        <div style="font-size:22px;font-weight:700;color:#e2ecf8;margin-bottom:4px">All Themes — Net Sentiment Scores</div>
        <div style="font-size:12px;color:#4a6080">NSS = (Positive − Negative) ÷ Total sentences × 100 · All 18 themes · Segmented by Setting, Specialty, Target</div>
    </div>''', unsafe_allow_html=True)

    with st.spinner("Running sentiment analysis across all 18 themes..."):
        all_theme_results = {}
        for theme in THEMES:
            sa = run_sentiment_analysis(theme, full_df)
            if sa["total_sentences"] > 0:
                all_theme_results[theme] = sa

    if not all_theme_results:
        st.warning("No sentiment data found.")
        return

    T = len(full_df)

    # ── Overall NSS per theme ─────────────────────────────────────────────
    st.markdown('<div class="sec-lbl">NET SENTIMENT SCORE — ALL THEMES</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:14px">Ranked by NSS. Green = net positive · Red = net negative · Amber = balanced. Click a theme in the sidebar to see full detail.</div>', unsafe_allow_html=True)

    # Sort by NSS descending
    theme_nss = []
    for theme, sa in all_theme_results.items():
        c = sa["counts"]
        denom = c["POSITIVE"] + c["NEGATIVE"] + c["NEUTRAL"]
        nss = round((c["POSITIVE"] - c["NEGATIVE"]) / denom * 100) if denom > 0 else 0
        theme_nss.append((theme, nss, c["POSITIVE"], c["NEGATIVE"], c["NEUTRAL"], c["MIXED"], sa["total_sentences"], sa["n_respondents"]))
    theme_nss.sort(key=lambda x: -x[1])

    st.markdown('''<div style="background:#0d1420;border:1px solid #1a2640;border-radius:10px;overflow:hidden;margin-bottom:20px">
        <div style="display:grid;grid-template-columns:180px 1fr 60px 60px 60px 70px 70px 90px;font-size:10px;font-weight:600;color:#4a6080;letter-spacing:0.5px;text-transform:uppercase;padding:8px 14px;background:#0a1220;border-bottom:1px solid #1a2640">
            <span>Theme</span><span>NSS bar</span>
            <span style="text-align:center">HCPs</span>
            <span style="text-align:center">Pos</span>
            <span style="text-align:center">Neg</span>
            <span style="text-align:center">Neutral</span>
            <span style="text-align:center">Sentences</span>
            <span style="text-align:center">NSS</span>
        </div>''', unsafe_allow_html=True)

    for theme, nss, pos, neg, neu, mix, total_s, n_resp in theme_nss:
        color = TC.get(theme, "#4a6080")
        nss_c = "#10b981" if nss > 10 else "#ef4444" if nss < -10 else "#f59e0b"
        nss_label = "Net +" if nss > 10 else "Net −" if nss < -10 else "Balanced"
        bar_w = min(abs(nss), 100)
        bar_color = nss_c
        st.markdown(f'''<div style="display:grid;grid-template-columns:180px 1fr 60px 60px 60px 70px 70px 90px;align-items:center;padding:7px 14px;border-bottom:1px solid #0f1823">
            <div style="font-size:12px;color:#e2ecf8;font-weight:500">
                <span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:{color};margin-right:6px"></span>{theme}
            </div>
            <div style="display:flex;height:10px;border-radius:3px;overflow:hidden;margin-right:8px;background:#1e293b">
                <div style="width:{bar_w}%;background:{bar_color};border-radius:3px"></div>
            </div>
            <div style="text-align:center;font-size:12px;color:#e2ecf8;font-weight:500">{n_resp}</div>
            <div style="text-align:center;font-size:12px;color:#10b981">{pos}</div>
            <div style="text-align:center;font-size:12px;color:{"#ef4444" if neg > 0 else "#4a6080"}">{neg if neg > 0 else "—"}</div>
            <div style="text-align:center;font-size:12px;color:#4a6080">{neu}</div>
            <div style="text-align:center;font-size:12px;color:#64748b">{total_s}</div>
            <div style="text-align:center">
                <span style="font-family:'IBM Plex Mono',monospace;font-size:14px;font-weight:700;color:{nss_c}">{nss:+d}</span>
                <div style="font-size:9px;color:{nss_c}">{nss_label}</div>
            </div>
        </div>''', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    # ── NSS by segment across all themes ─────────────────────────────────
    st.markdown('<div class="sec-lbl" style="margin-top:20px">NSS BY SEGMENT — ALL THEMES</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:14px">How sentiment compares across segments for every theme. No quotes — stats only.</div>', unsafe_allow_html=True)

    for seg_key, seg_label in [("setting","Practice Setting"),("specialty","Specialty"),("target","Target Type")]:
        st.markdown(f'<div class="card"><div class="sec-lbl">{seg_label}</div>', unsafe_allow_html=True)

        # Get all unique values for this segment
        seg_vals = sorted(full_df[seg_key].dropna().unique().tolist())
        if not seg_vals: continue

        # Header
        col_w = f"160px " + " ".join([f"80px"] * len(seg_vals))
        header_html = f'<div style="display:grid;grid-template-columns:{col_w};font-size:10px;font-weight:600;color:#4a6080;letter-spacing:0.5px;text-transform:uppercase;padding:6px 8px;background:#0a1220;border-radius:6px;margin-bottom:4px"><span>Theme</span>'
        for v in seg_vals:
            header_html += f'<span style="text-align:center">{v[:12]}</span>'
        header_html += '</div>'
        st.markdown(header_html, unsafe_allow_html=True)

        for theme, nss_overall, *_ in theme_nss:
            sa = all_theme_results.get(theme)
            if not sa: continue
            seg_data = sa["seg_sentiment"].get(seg_key, {})
            row_html = f'<div style="display:grid;grid-template-columns:{col_w};align-items:center;padding:5px 8px;border-bottom:1px solid #0f1823"><span style="font-size:11px;color:#e2ecf8">{theme}</span>'
            for v in seg_vals:
                d = seg_data.get(v)
                if d:
                    denom = d["pos"] + d["neg"] + d.get("neutral", 0)
                    seg_nss = round((d["pos"] - d["neg"]) / denom * 100) if denom > 0 else 0
                    nc = "#10b981" if seg_nss > 10 else "#ef4444" if seg_nss < -10 else "#f59e0b"
                    warn = "⚠" if d.get("n_resp", 0) < 5 else ""
                    row_html += f'<div style="text-align:center"><span style="font-family:monospace;font-size:12px;font-weight:700;color:{nc}">{seg_nss:+d}{warn}</span></div>'
                else:
                    row_html += '<div style="text-align:center;color:#2a3a55">—</div>'
            row_html += '</div>'
            st.markdown(row_html, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Download as Excel ─────────────────────────────────────────────────
    st.markdown('<div class="sec-lbl" style="margin-top:20px">⬇ DOWNLOAD SENTIMENT RESULTS</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:12px">Full sentiment data including all classified sentences, HCP-level NSS, and segment breakdowns — formatted Excel workbook.</div>', unsafe_allow_html=True)

    if st.button("⬇ Generate & Download Sentiment Excel", key="dl_overall_excel"):
        wb = Workbook()

        def hcell(ws, r, c, v, bg="0B1F3A", fg="FFFFFF", bold=True, sz=10, align="center"):
            cell = ws.cell(r, c, v)
            cell.font = Font(name="Arial", bold=bold, color=fg, size=sz)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            return cell

        def dcell(ws, r, c, v, fg="1E293B", bold=False, sz=10, align="left", bg=None):
            cell = ws.cell(r, c, v)
            cell.font = Font(name="Arial", bold=bold, color=fg, size=sz)
            if bg: cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            return cell

        # Sheet 1: Summary — all themes NSS
        ws1 = wb.active; ws1.title = "Summary — All Themes"
        ws1.sheet_view.showGridLines = False
        headers = ["Theme","HCPs Mentioning","Total Sentences","Positive","Negative","Mixed","Neutral","NSS","NSS Label"]
        for col, h in enumerate(headers, 1):
            hcell(ws1, 1, col, h)
        for i, (theme, nss, pos, neg, neu, mix, total_s, n_resp) in enumerate(theme_nss, 2):
            nss_label = "Net positive" if nss > 10 else "Net negative" if nss < -10 else "Balanced"
            bg = "ECFDF5" if nss > 10 else "FEF2F2" if nss < -10 else "FFFBEB"
            fg_nss = "065F46" if nss > 10 else "7F1D1D" if nss < -10 else "92400E"
            for col, v in enumerate([theme, n_resp, total_s, pos, neg, mix, neu], 1):
                dcell(ws1, i, col, v, align="center")
            dcell(ws1, i, 8, f"{nss:+d}", fg_nss, True, 10, "center", bg)
            dcell(ws1, i, 9, nss_label, fg_nss, False, 10, "center", bg)
            ws1.row_dimensions[i].height = 20
        for col, w in zip(range(1,10), [22,14,14,10,10,10,10,10,14]):
            ws1.column_dimensions[__import__("openpyxl").utils.get_column_letter(col)].width = w

        # Sheet 2: All classified sentences
        ws2 = wb.create_sheet("All Sentences")
        ws2.sheet_view.showGridLines = False
        sent_headers = ["Theme","HCP_ID","Practice Setting","Specialty","Target","Sentiment","Confidence","Positive Triggers","Negative Triggers","Full Sentence"]
        for col, h in enumerate(sent_headers, 1):
            hcell(ws2, 1, col, h, "0D9488")
        row_idx = 2
        sent_colors = {"POSITIVE":"ECFDF5","NEGATIVE":"FEF2F2","MIXED":"FFFBEB","NEUTRAL":None}
        sent_fg = {"POSITIVE":"065F46","NEGATIVE":"7F1D1D","MIXED":"92400E","NEUTRAL":"334155"}
        for theme, sa in all_theme_results.items():
            for r in sa["all_results"]:
                bg = sent_colors.get(r["sentiment"])
                fg = sent_fg.get(r["sentiment"],"334155")
                vals = [theme, r["id"], r["setting"], r["specialty"], r["target"],
                        r["sentiment"], r["confidence"],
                        ", ".join(r["pos_triggers"]), ", ".join(r["neg_triggers"]),
                        r["sentence"]]
                for col, v in enumerate(vals, 1):
                    dcell(ws2, row_idx, col, v,
                          fg if col==6 else "1E293B",
                          col==6, 9, "left", bg if col==6 else None)
                ws2.row_dimensions[row_idx].height = 30
                row_idx += 1
        for col, w in zip(range(1,11), [22,10,16,14,12,12,11,28,22,50]):
            ws2.column_dimensions[__import__("openpyxl").utils.get_column_letter(col)].width = w

        # Sheet 3: NSS by segment
        ws3 = wb.create_sheet("NSS by Segment")
        ws3.sheet_view.showGridLines = False
        hcell(ws3, 1, 1, "NSS BY SEGMENT — ALL THEMES", "0B1F3A", merge_to=None)
        r = 2
        for seg_key, seg_label in [("setting","Practice Setting"),("specialty","Specialty"),("target","Target Type")]:
            seg_vals = sorted(full_df[seg_key].dropna().unique().tolist())
            hcell(ws3, r, 1, seg_label, "0D9488"); r+=1
            hcell(ws3, r, 1, "Theme", "1E3A5F")
            for ci, v in enumerate(seg_vals, 2):
                hcell(ws3, r, ci, v[:20], "1E3A5F")
            r+=1
            for theme, *_ in theme_nss:
                sa = all_theme_results.get(theme)
                if not sa: continue
                dcell(ws3, r, 1, theme, "0B1F3A", True, 10, "left")
                for ci, v in enumerate(seg_vals, 2):
                    d = sa["seg_sentiment"].get(seg_key, {}).get(v)
                    if d:
                        denom2 = d["pos"]+d["neg"]+d.get("neutral",0)
                        sn = round((d["pos"]-d["neg"])/denom2*100) if denom2 else 0
                        nc = "065F46" if sn>10 else "7F1D1D" if sn<-10 else "92400E"
                        bg2 = "ECFDF5" if sn>10 else "FEF2F2" if sn<-10 else "FFFBEB"
                        dcell(ws3, r, ci, f"{sn:+d}", nc, True, 10, "center", bg2)
                    else:
                        dcell(ws3, r, ci, "—", "CBD5E1", False, 10, "center")
                ws3.row_dimensions[r].height = 20; r+=1
            r+=1
        ws3.column_dimensions["A"].width = 22
        for ci in range(2, 10):
            ws3.column_dimensions[__import__("openpyxl").utils.get_column_letter(ci)].width = 16

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.download_button(
            "⬇ Download Sentiment_Results.xlsx",
            buf.getvalue(),
            "Sentiment_Results.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


def render_theme_sentiment(theme, full_df):
    """
    Sentiment view for a single theme — stats + NSS only.
    No quotes on screen. Download only.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    with st.spinner(f"Running sentiment analysis for {theme}..."):
        sa = run_sentiment_analysis(theme, full_df)

    T_sents = sa["total_sentences"]
    if T_sents == 0:
        st.warning(f"No sentences referencing {theme} found.")
        return

    color = TC.get(theme, "#3b6ef7")
    counts = sa["counts"]
    pos_n = counts["POSITIVE"]; neg_n = counts["NEGATIVE"]
    mix_n = counts["MIXED"];    neu_n = counts["NEUTRAL"]
    denom = pos_n + neg_n + neu_n
    nss = round((pos_n - neg_n) / denom * 100, 1) if denom > 0 else 0
    nss_color = "#10b981" if nss > 10 else "#ef4444" if nss < -10 else "#f59e0b"
    nss_label = "Net positive" if nss > 10 else "Net negative" if nss < -10 else "Balanced"

    # Theme header
    st.markdown(f'''<div style="background:#0d1420;border:1px solid #1a2640;border-radius:12px;padding:16px 20px;margin-bottom:16px;border-left:4px solid {color}">
        <div style="font-size:10px;letter-spacing:2px;text-transform:uppercase;color:{color};font-weight:600;margin-bottom:4px;font-family:'IBM Plex Mono',monospace">SENTIMENT ANALYSIS</div>
        <div style="font-size:22px;font-weight:700;color:#e2ecf8">{theme}</div>
        <div style="font-size:11px;color:#4a6080;margin-top:4px">{T_sents} sentences from {sa["n_respondents"]} respondents · {len(full_df)} total in dataset</div>
    </div>''', unsafe_allow_html=True)

    # NSS score card
    nss_interp = (f"Doctors broadly express favourable sentiment on {theme}."
                  if nss > 10 else
                  f"Concern or hesitation language outweighs positive on {theme}."
                  if nss < -10 else
                  f"Sentiment on {theme} is balanced — positive and negative roughly equal.")

    st.markdown(f'''<div class="card" style="border-left:4px solid {nss_color};margin-bottom:16px">
        <div style="display:flex;align-items:center;gap:24px;flex-wrap:wrap">
            <div style="text-align:center;min-width:110px">
                <div style="font-size:10px;color:{nss_color};font-weight:600;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:4px">Net Sentiment Score</div>
                <div style="font-family:'IBM Plex Mono',monospace;font-size:44px;font-weight:700;color:{nss_color};line-height:1">{nss:+.1f}</div>
                <div style="font-size:11px;color:{nss_color};margin-top:2px">{nss_label}</div>
            </div>
            <div style="flex:1;min-width:200px">
                <div style="font-size:13px;color:#c8d4e8;line-height:1.6;margin-bottom:8px">{nss_interp}</div>
                <div style="font-size:11px;color:#4a6080">Formula: (Positive − Negative) ÷ (Pos + Neg + Neutral) × 100<br>
                ({pos_n} − {neg_n}) ÷ ({pos_n} + {neg_n} + {neu_n}) × 100 = <b style="color:{nss_color}">{nss:+.1f}</b></div>
            </div>
        </div>
    </div>''', unsafe_allow_html=True)

    # Stat cards
    c1,c2,c3,c4 = st.columns(4)
    for col, (label, count, col_c, desc) in zip([c1,c2,c3,c4], [
        ("Positive", pos_n, "#10b981", "Favourable language"),
        ("Negative", neg_n, "#ef4444", "Concern / hesitation"),
        ("Mixed",    mix_n, "#f59e0b", "Both signals present"),
        ("Neutral",  neu_n, "#64748b", "Clinical / factual"),
    ]):
        pct = round(count/T_sents*100) if T_sents else 0
        col.markdown(f'''<div class="card" style="text-align:center;border-left:3px solid {col_c}">
            <div style="font-size:10px;color:{col_c};font-weight:600;margin-bottom:4px;text-transform:uppercase">{label}</div>
            <div style="font-family:'IBM Plex Mono',monospace;font-size:28px;font-weight:600;color:{col_c}">{count}</div>
            <div style="font-size:13px;color:#e2ecf8;margin-top:2px">{pct}%</div>
            <div style="font-size:9px;color:#4a6080;margin-top:4px">{desc}</div>
        </div>''', unsafe_allow_html=True)

    # Sentiment bar
    pos_pct = round(pos_n/T_sents*100) if T_sents else 0
    neg_pct = round(neg_n/T_sents*100) if T_sents else 0
    mix_pct = round(mix_n/T_sents*100) if T_sents else 0
    neu_pct = 100 - pos_pct - neg_pct - mix_pct
    st.markdown(f'''<div style="margin:14px 0">
        <div style="font-size:11px;color:#4a6080;margin-bottom:5px">Sentence distribution — {T_sents} total</div>
        <div style="display:flex;height:16px;border-radius:4px;overflow:hidden;gap:2px">
            <div style="width:{pos_pct}%;background:#10b981"></div>
            <div style="width:{neg_pct}%;background:#ef4444"></div>
            <div style="width:{mix_pct}%;background:#f59e0b"></div>
            <div style="flex:1;background:#1e293b"></div>
        </div>
        <div style="display:flex;gap:14px;margin-top:5px;font-size:11px">
            <span style="color:#10b981">■ Positive {pos_pct}%</span>
            <span style="color:#ef4444">■ Negative {neg_pct}%</span>
            <span style="color:#f59e0b">■ Mixed {mix_pct}%</span>
            <span style="color:#64748b">■ Neutral {neu_pct}%</span>
        </div>
    </div>''', unsafe_allow_html=True)

    # Segment breakdown — stats + NSS, no quotes
    st.markdown('<div class="sec-lbl" style="margin-top:16px">SENTIMENT BY SEGMENT</div>', unsafe_allow_html=True)

    for seg_key, seg_label in [("setting","Practice Setting"),("specialty","Specialty"),("target","Target Type")]:
        seg_data = sa["seg_sentiment"].get(seg_key, {})
        if not seg_data: continue
        st.markdown(f'''<div class="card" style="margin-bottom:14px">
            <div class="sec-lbl" style="margin-bottom:12px">{seg_label}</div>
            <div style="display:grid;grid-template-columns:160px 1fr 60px 60px 60px 70px 70px 90px;font-size:10px;font-weight:600;color:#4a6080;letter-spacing:0.5px;text-transform:uppercase;padding:4px 0;border-bottom:1px solid #1a2640;margin-bottom:6px">
                <span>Segment</span><span>Sentiment bar</span>
                <span style="text-align:center">HCPs</span>
                <span style="text-align:center">Pos</span>
                <span style="text-align:center">Neg</span>
                <span style="text-align:center">Neutral</span>
                <span style="text-align:center">Sentences</span>
                <span style="text-align:center">NSS</span>
            </div>''', unsafe_allow_html=True)

        for val, d in sorted(seg_data.items(), key=lambda x: -(x[1]["pos"]+x[1]["neg"])):
            vt = d["total"]
            if vt == 0: continue
            pos2=d["pos"]; neg2=d["neg"]; mix2=d.get("mixed",0); neu2=d.get("neutral",0)
            n_resp2=d.get("n_resp",0)
            p=round(pos2/vt*100); n=round(neg2/vt*100); m=round(mix2/vt*100)
            denom2=pos2+neg2+neu2
            sn=round((pos2-neg2)/denom2*100) if denom2 else 0
            nc="#10b981" if sn>10 else "#ef4444" if sn<-10 else "#f59e0b"
            nl="Net +" if sn>10 else "Net −" if sn<-10 else "Balanced"
            warn="⚠" if n_resp2<5 or vt<10 else ""
            pos_s=f"{p}%" if p>0 else "—"; neg_s=f"{n}%" if n>0 else "—"
            st.markdown(f'''<div style="display:grid;grid-template-columns:160px 1fr 60px 60px 60px 70px 70px 90px;align-items:center;padding:7px 0;border-bottom:1px solid #0f1823">
                <div style="font-size:12px;color:#e2ecf8;font-weight:500">{val[:22]}{warn}</div>
                <div style="display:flex;height:10px;border-radius:3px;overflow:hidden;gap:2px;margin-right:8px;background:#1e293b">
                    <div style="width:{p}%;background:#10b981"></div>
                    <div style="width:{n}%;background:#ef4444"></div>
                    <div style="width:{m}%;background:#f59e0b"></div>
                </div>
                <div style="text-align:center"><div style="font-size:13px;color:#e2ecf8;font-weight:500">{n_resp2}</div><div style="font-size:9px;color:#4a6080">HCPs</div></div>
                <div style="text-align:center;font-size:12px;color:#10b981">{pos_s}</div>
                <div style="text-align:center;font-size:12px;color:{"#ef4444" if n>0 else "#4a6080"}">{neg_s}</div>
                <div style="text-align:center;font-size:12px;color:#4a6080">{neu2}</div>
                <div style="text-align:center"><div style="font-size:13px;color:#64748b">{vt}</div><div style="font-size:9px;color:#4a6080">sent.</div></div>
                <div style="text-align:center">
                    <span style="font-family:'IBM Plex Mono',monospace;font-size:14px;font-weight:700;color:{nc}">{sn:+d}</span>
                    <div style="font-size:9px;color:{nc}">{nl}</div>
                </div>
            </div>''', unsafe_allow_html=True)

        # Calculation workings
        st.markdown('<div style="margin-top:10px;border-top:1px solid #1a2640;padding-top:10px">', unsafe_allow_html=True)
        st.markdown('<div style="font-size:10px;font-weight:600;color:#4a6080;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">NSS calculation workings</div>', unsafe_allow_html=True)
        st.markdown('''<div style="display:grid;grid-template-columns:160px 50px 50px 50px 50px 1fr 80px;font-size:9px;font-weight:600;color:#4a6080;text-transform:uppercase;padding:4px 8px;background:#0a1220;border-radius:6px 6px 0 0">
            <span>Segment</span><span style="text-align:center">Pos</span><span style="text-align:center">Neg</span>
            <span style="text-align:center">Neutral</span><span style="text-align:center">Total</span>
            <span style="text-align:center">Formula</span><span style="text-align:center">NSS</span>
        </div>''', unsafe_allow_html=True)
        small_segs=[]
        for val, d in sorted(seg_data.items(), key=lambda x: -(x[1]["pos"]+x[1]["neg"])):
            vt=d["total"]
            if vt==0: continue
            pos2=d["pos"]; neg2=d["neg"]; neu2=d.get("neutral",0)
            denom2=pos2+neg2+neu2
            sn=round((pos2-neg2)/denom2*100) if denom2 else 0
            nc="#10b981" if sn>10 else "#ef4444" if sn<-10 else "#f59e0b"
            formula=f"({pos2}−{neg2})÷({pos2}+{neg2}+{neu2})×100 = {pos2-neg2}÷{denom2}×100"
            n_resp2=d.get("n_resp",0)
            if n_resp2<5 or vt<10: small_segs.append(val)
            st.markdown(f'''<div style="display:grid;grid-template-columns:160px 50px 50px 50px 50px 1fr 80px;font-size:11px;align-items:center;padding:5px 8px;border-bottom:1px solid #0f1823;background:#0d1420">
                <span style="color:#c8d4e8;font-weight:500">{val[:22]}</span>
                <span style="text-align:center;color:#10b981;font-family:'IBM Plex Mono',monospace">{pos2}</span>
                <span style="text-align:center;color:{"#ef4444" if neg2>0 else "#4a6080"};font-family:'IBM Plex Mono',monospace">{neg2}</span>
                <span style="text-align:center;color:#4a6080;font-family:'IBM Plex Mono',monospace">{neu2}</span>
                <span style="text-align:center;color:#64748b;font-family:'IBM Plex Mono',monospace">{denom2}</span>
                <span style="color:#64748b;font-size:10px;font-family:'IBM Plex Mono',monospace">{formula}</span>
                <span style="text-align:center;font-family:'IBM Plex Mono',monospace;font-weight:700;color:{nc}">{sn:+d}</span>
            </div>''', unsafe_allow_html=True)
        if small_segs:
            st.markdown(f'<div style="font-size:10px;color:#f59e0b;padding:6px 8px;background:#1a1200;border-radius:0 0 6px 6px">⚠ Small sample: {", ".join(small_segs[:3])} — treat directionally (n&lt;5 HCPs or &lt;10 sentences)</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div style="height:4px;background:#0a1220;border-radius:0 0 6px 6px"></div>', unsafe_allow_html=True)
        st.markdown('</div></div>', unsafe_allow_html=True)

    # NSS methodology note
    st.markdown(f'''<div class="card" style="border-left:3px solid #3b6ef7;margin-top:4px">
        <div class="sec-lbl">HOW NSS WAS CALCULATED</div>
        <div style="font-size:12px;color:#c8d4e8;line-height:1.75">
            Each SPEAKER_B response split into sentences → sentences referencing <b>{theme}</b> identified →
            each sentence classified Positive / Negative / Mixed / Neutral using a pharma-tuned lexicon
            with negation detection and context overrides →
            NSS = (Positive − Negative) ÷ (Pos + Neg + Neutral) × 100.
        </div>
        <div style="margin-top:8px;font-size:11px;color:#4a6080">
            Score ranges: <b style="color:#10b981">+10 to +100</b> = Net positive ·
            <b style="color:#f59e0b">-10 to +10</b> = Balanced ·
            <b style="color:#ef4444">-100 to -10</b> = Net negative
        </div>
    </div>''', unsafe_allow_html=True)

    # Download button — Excel
    st.markdown('<div class="sec-lbl" style="margin-top:20px">⬇ DOWNLOAD SENTIMENT DATA</div>', unsafe_allow_html=True)
    st.markdown(f'<div style="font-size:11px;color:#4a6080;margin-bottom:10px">Classified sentences, HCP-level NSS, and segment breakdown for {theme}. No quotes shown on screen — all in the Excel.</div>', unsafe_allow_html=True)

    if st.button(f"⬇ Generate & Download {theme} Sentiment Excel", key=f"dl_theme_{theme}"):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        def hc(ws,r,c,v,bg="0B1F3A",fg="FFFFFF",bold=True,sz=10):
            cell=ws.cell(r,c,v); cell.font=Font(name="Arial",bold=bold,color=fg,size=sz)
            cell.fill=PatternFill("solid",fgColor=bg); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            return cell
        def dc(ws,r,c,v,fg="1E293B",bold=False,sz=10,align="left",bg=None):
            cell=ws.cell(r,c,v); cell.font=Font(name="Arial",bold=bold,color=fg,size=sz)
            if bg: cell.fill=PatternFill("solid",fgColor=bg)
            cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
            return cell

        sent_colors={"POSITIVE":"ECFDF5","NEGATIVE":"FEF2F2","MIXED":"FFFBEB","NEUTRAL":None}
        sent_fg={"POSITIVE":"065F46","NEGATIVE":"7F1D1D","MIXED":"92400E","NEUTRAL":"334155"}

        # Sheet 1: All sentences
        ws1=wb.active; ws1.title="All Sentences"
        ws1.sheet_view.showGridLines=False
        for col,h in enumerate(["HCP_ID","Practice Setting","Specialty","Target","Sentiment","Confidence","Positive Triggers","Negative Triggers","Full Sentence"],1):
            hc(ws1,1,col,h,"0D9488")
        for ri,r in enumerate(sa["all_results"],2):
            bg=sent_colors.get(r["sentiment"]); fg2=sent_fg.get(r["sentiment"],"334155")
            vals=[r["id"],r["setting"],r["specialty"],r["target"],r["sentiment"],r["confidence"],", ".join(r["pos_triggers"]),", ".join(r["neg_triggers"]),r["sentence"]]
            for col,v in enumerate(vals,1):
                dc(ws1,ri,col,v,fg2 if col==5 else "1E293B",col==5,9,"left",bg if col==5 else None)
            ws1.row_dimensions[ri].height=30
        for col,w in zip(range(1,10),[10,16,14,12,12,11,28,22,55]):
            ws1.column_dimensions[__import__("openpyxl").utils.get_column_letter(col)].width=w

        # Sheet 2: HCP NSS summary
        ws2=wb.create_sheet("HCP NSS Summary")
        ws2.sheet_view.showGridLines=False
        hcp_data={}
        for r in sa["all_results"]:
            hid=r["id"]
            if hid not in hcp_data:
                hcp_data[hid]={"id":hid,"setting":r["setting"],"specialty":r["specialty"],"target":r["target"],"pos":0,"neg":0,"mix":0,"neu":0}
            hcp_data[hid][{"POSITIVE":"pos","NEGATIVE":"neg","MIXED":"mix","NEUTRAL":"neu"}.get(r["sentiment"],"neu")]+=1
        for col,h in enumerate(["HCP_ID","Practice Setting","Specialty","Target","Positive","Negative","Mixed","Neutral","Total","NSS","NSS Label"],1):
            hc(ws2,1,col,h,"0D9488")
        hcp_list=[]
        for h in hcp_data.values():
            tot=h["pos"]+h["neg"]+h["mix"]+h["neu"]
            denom2=h["pos"]+h["neg"]+h["neu"]
            hn=round((h["pos"]-h["neg"])/denom2*100,1) if denom2 else 0
            hl="Net positive" if hn>10 else "Net negative" if hn<-10 else "Balanced"
            hcp_list.append((h,hn,hl,tot))
        hcp_list.sort(key=lambda x:-x[1])
        for ri,(h,hn,hl,tot) in enumerate(hcp_list,2):
            bg2="ECFDF5" if hn>10 else "FEF2F2" if hn<-10 else "FFFBEB"
            fg3="065F46" if hn>10 else "7F1D1D" if hn<-10 else "92400E"
            for col,v in enumerate([h["id"],h["setting"],h["specialty"],h["target"],h["pos"],h["neg"],h["mix"],h["neu"],tot],1):
                dc(ws2,ri,col,v,"1E293B",False,10,"center")
            dc(ws2,ri,10,f"{hn:+.1f}",fg3,True,11,"center",bg2)
            dc(ws2,ri,11,hl,fg3,False,10,"center",bg2)
            ws2.row_dimensions[ri].height=20
        for col,w in zip(range(1,12),[12,18,14,12,10,10,10,10,10,12,14]):
            ws2.column_dimensions[__import__("openpyxl").utils.get_column_letter(col)].width=w

        # Sheet 3: Segment NSS
        ws3=wb.create_sheet("Segment NSS")
        ws3.sheet_view.showGridLines=False
        r3=1
        for seg_key,seg_label in [("setting","Practice Setting"),("specialty","Specialty"),("target","Target Type")]:
            hc(ws3,r3,1,seg_label,"1E3A5F"); r3+=1
            for col,h in enumerate(["Segment","HCPs","Total Sentences","Positive","Negative","Neutral","NSS","NSS Label"],1):
                hc(ws3,r3,col,h,"0D9488"); 
            r3+=1
            for val,d in sorted(sa["seg_sentiment"].get(seg_key,{}).items(),key=lambda x:-(x[1]["pos"]+x[1]["neg"])):
                vt=d["total"]; pos2=d["pos"]; neg2=d["neg"]; neu2=d.get("neutral",0)
                denom2=pos2+neg2+neu2; sn=round((pos2-neg2)/denom2*100,1) if denom2 else 0
                sl2="Net positive" if sn>10 else "Net negative" if sn<-10 else "Balanced"
                bg3="ECFDF5" if sn>10 else "FEF2F2" if sn<-10 else "FFFBEB"
                fg4="065F46" if sn>10 else "7F1D1D" if sn<-10 else "92400E"
                for col,v in enumerate([val,d.get("n_resp",0),vt,pos2,neg2,neu2],1):
                    dc(ws3,r3,col,v,"1E293B",False,10,"center")
                dc(ws3,r3,7,f"{sn:+.1f}",fg4,True,11,"center",bg3)
                dc(ws3,r3,8,sl2,fg4,False,10,"center",bg3)
                ws3.row_dimensions[r3].height=20; r3+=1
            r3+=1
        for col,w in zip(range(1,9),[22,10,14,10,10,10,12,14]):
            ws3.column_dimensions[__import__("openpyxl").utils.get_column_letter(col)].width=w

        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        st.download_button(
            f"⬇ Download {theme}_Sentiment.xlsx",
            buf.getvalue(),
            f"{theme.replace('/','_')}_Sentiment.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )



def render_dashboard(theme, full_df):
    pats = THEMES.get(theme, [theme.lower()])
    theme_df = full_df[full_df["text_lower"].apply(lambda x: any(p in x for p in pats))]
    T = len(full_df); n = len(theme_df); color = TC.get(theme, "#3b6ef7")
    seg_data = build_segment_breakdown(theme, full_df)

    st.markdown(f"""<div style="background:#0d1420;border:1px solid #1a2640;border-radius:14px;padding:20px 24px;margin-bottom:20px;border-left:4px solid {color}">
        <div style="font-size:10px;letter-spacing:2px;text-transform:uppercase;color:{color};font-weight:600;margin-bottom:6px;font-family:'IBM Plex Mono',monospace">THEME DASHBOARD</div>
        <div style="font-size:26px;font-weight:700;color:#e2ecf8;margin-bottom:4px">{theme}</div>
        <div style="font-size:13px;color:#4a6080">Based on {T} uploaded respondents · All data from your Excel · No inference</div>
    </div>""", unsafe_allow_html=True)

    if n == 0:
        st.warning(f"No respondents mentioned {theme} in the uploaded data.")
        return

    tab1, tab2, tab3 = st.tabs(["📋  Dashboard Summary", "💬  Quotes & Evidence", "⬇  Download"])

    with tab1:
        # Stat row
        co = {}
        for t, p2 in THEMES.items():
            if t == theme: continue
            c = int(theme_df["text_lower"].apply(lambda x: any(p in x for p in p2)).sum())
            if c: co[t] = c
        top_co = max(co, key=co.get) if co else "—"
        solo = sum(1 for _,row in theme_df.iterrows() if sum(1 for t,p2 in THEMES.items() if t!=theme and any(p in row["text_lower"] for p in p2)) <= 1)
        verdict = "Standalone" if solo > n//2 else "Complex"
        vc = "#10b981" if verdict == "Standalone" else "#ef4444"

        c1,c2,c3,c4 = st.columns(4)
        c1.markdown(f'<div class="card" style="text-align:center;border-left:3px solid {color}"><div class="stat-num" style="color:{color}">{n}</div><div class="stat-lbl">respondents</div></div>', unsafe_allow_html=True)
        c2.markdown(f'<div class="card" style="text-align:center"><div class="stat-num">{round(n/T*100)}%</div><div class="stat-lbl">of total sample</div></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="card" style="text-align:center"><div class="stat-num" style="font-size:16px;line-height:1.4">{top_co}</div><div class="stat-lbl">top co-theme</div></div>', unsafe_allow_html=True)
        c4.markdown(f'<div class="card" style="text-align:center"><div class="stat-num" style="color:{vc};font-size:20px">{verdict}</div><div class="stat-lbl">driver type</div></div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # Executive summary
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown(f'<div class="sec-lbl">5-POINT EXECUTIVE SUMMARY</div>', unsafe_allow_html=True)
        st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:16px">Synthesised from uploaded data only. Every point is backed by counts from your Excel.</div>', unsafe_allow_html=True)
        for i, pt in enumerate(build_exec_summary(theme, theme_df, full_df, seg_data), 1):
            st.markdown(f"""<div style="display:flex;gap:14px;margin-bottom:14px;padding-bottom:14px;border-bottom:1px solid #1a2640;align-items:flex-start">
                <div style="font-family:'IBM Plex Mono',monospace;font-size:18px;color:{color};font-weight:700;min-width:28px;line-height:1.4">{i}</div>
                <div style="font-size:14px;color:#c8d4e8;line-height:1.75">{pt}</div>
            </div>""", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        # Segment breakdowns
        seg_cols_ui = st.columns(3)
        for col_w, (seg_col, label) in zip(seg_cols_ui, [("setting","BY PRACTICE SETTING"),("specialty","BY SPECIALTY"),("target","BY TARGET TYPE")]):
            bd = seg_data.get(seg_col, {})
            if not bd: continue
            with col_w:
                st.markdown(f'<div class="card"><div class="sec-lbl">{label}</div>', unsafe_allow_html=True)
                for val, dp in sorted(bd.items(), key=lambda x: -x[1]["pct"]):
                    st.markdown(f"""<div style="margin-bottom:10px">
                        <div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:3px">
                            <span style="color:#e2ecf8;font-weight:500">{val}</span>
                            <span style="font-family:'IBM Plex Mono',monospace;color:{color}">{dp["n"]}/{dp["total"]} ({dp["pct"]}%)</span>
                        </div>
                        <div class="bar-trk"><div class="bar-fill" style="width:{dp["pct"]}%;background:{color}"></div></div>
                    </div>""", unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)

        # Co-theme chart
        st.markdown(f'<div class="card"><div class="sec-lbl">MOST ASSOCIATED THEMES</div>', unsafe_allow_html=True)
        st.markdown(f'<div style="font-size:11px;color:#4a6080;margin-bottom:12px">Among {n} respondents who mentioned {theme}, these other themes also appeared.</div>', unsafe_allow_html=True)
        for t_co, cnt_co in sorted(co.items(), key=lambda x: -x[1])[:10]:
            if cnt_co: bar_html(t_co, cnt_co, n, TC.get(t_co, "#4a6080"))
        st.markdown('</div>', unsafe_allow_html=True)

    with tab2:
        all_settings = ["All"] + sorted(theme_df["setting"].dropna().unique().tolist())
        filter_s = st.selectbox("Filter by Setting", all_settings, key=f"ds_{theme}")
        disp = theme_df if filter_s == "All" else theme_df[theme_df["setting"] == filter_s]
        st.markdown(f'<div style="font-size:11px;color:#4a6080;margin-bottom:12px">Showing {len(disp)} respondents. Full verbatim text, all themes colour-highlighted.</div>', unsafe_allow_html=True)
        for _, row in disp.iterrows():
            quote_card(row, focus=[theme])

    with tab3:
        export_rows = []
        for _, row in theme_df.iterrows():
            others = [t for t,p2 in THEMES.items() if t!=theme and any(p in row["text_lower"] for p in p2)]
            export_rows.append({"ID":row["id"],"Setting":row["setting"],"Specialty":row["specialty"],"Target":row.get("target",""),
                                 f"Mentioned {theme}":"Yes","Other Themes":", ".join(others[:8]),
                                 "N Other Themes":len(others),"Driver Type":"Standalone" if len(others)<=1 else "Complex" if len(others)>=4 else "Moderate",
                                 "Full Response":row["text"]})
        dfe = pd.DataFrame(export_rows)
        preview_cols = [c for c in ["ID","Setting","Specialty","Target","Driver Type","N Other Themes"] if c in dfe.columns]
        st.dataframe(dfe[preview_cols], hide_index=True, use_container_width=True)
        c1,c2 = st.columns(2)
        with c1:
            st.download_button(f"⬇ {theme} full responses CSV", dfe.to_csv(index=False).encode(), f"{theme.replace('/','_')}_responses.csv", "text/csv")
        with c2:
            sum_rows = [{"Segment Type":lbl,"Segment Value":val,"N Mentioned":dp["n"],"Total in Group":dp["total"],"% Mentioned":f"{dp['pct']}%"}
                        for seg_col,lbl in [("setting","Setting"),("specialty","Specialty"),("target","Target")]
                        for val,dp in seg_data.get(seg_col,{}).items()]
            st.download_button(f"⬇ {theme} segment summary CSV", pd.DataFrame(sum_rows).to_csv(index=False).encode(), f"{theme.replace('/','_')}_segments.csv", "text/csv")


# ── QUESTION BANK ─────────────────────────────────────────────────────────────
QB = {
    "📊 Frequency": ["How many HCPs mentioned PFS?","How many discussed cost or insurance?","How often was quality of life mentioned?","How many cited the Indigo trial?","What percentage mentioned seizures?","How many brought up radiation delay?","How many discussed oral administration?","How many mentioned NCCN guidelines?","How many raised fertility concerns?","How often was performance status cited?"],
    "⚖️ Comparison": ["Community vs academic on PFS — what's the difference?","Do community doctors mention cost more than academic?","Academic vs community on radiation delay","Did academic HCPs cite Indigo trial more than community?","Community vs academic on seizures","On TL vs Off TL — what's different?","What do Neuro-Onc vs MedOnc say about PFS?","HemeOnc vs MedOnc on cost barriers","Community vs academic on quality of life","Community vs academic on patient assistance"],
    "🎯 Drivers": ["What is the top driver for prescribing?","What driver is most tied to PFS?","What are the main barriers to prescribing?","Why do doctors delay radiation?","What motivates HCPs to choose oral therapy?","What are the key clinical rationale drivers?","What are the main cost-related barriers?","What stops community doctors from prescribing?","What is the primary driver for insurance navigation?","What are the main reasons for non-prescribing?"],
    "🔗 Co-occurrence": ["Did doctors talk about PFS and performance status hand in hand?","Did PFS and quality of life go together?","Did cost and patient assistance come up together?","When seizures were mentioned, was quality of life also discussed?","Did oral administration and convenience go hand in hand?","Did Indigo trial and PFS get mentioned together?","When NCCN was mentioned, was insurance also discussed?","Did radiation delay and cognition come up together?","Did PFS and OS get mentioned alongside each other?","Did safety and tolerability go hand in hand?"],
    "🧩 Theme Clustering": ["What other drivers were tagged to PFS?","What themes cluster around cost?","What travels with quality of life discussions?","What other topics come up with seizure mentions?","What themes cluster around the Indigo trial?","What goes with oral administration mentions?","What other factors travel with radiation delay?","What themes come up alongside NCCN guidelines?","What clusters around efficacy discussions?","What other themes appear with OS mentions?"],
    "🎯 Driver Complexity": ["Was PFS a straightforward driver or always complex?","Was oral administration a standalone driver?","Was cost a simple or complex barrier?","Was quality of life mentioned standalone or bundled?","Was efficacy a simple or entangled driver?","Was seizure control a standalone concern?","Was radiation delay a straightforward benefit?","Was the Indigo trial cited alone or bundled?","Was NCCN a standalone factor or always with others?","Was tolerability a simple or complex driver?"],
    "💬 Full Responses": ["Show me full responses about cost and insurance","Give me full quotes about PFS","Show full responses from community HCPs about barriers","Show full transcripts about quality of life","Show full responses mentioning seizures","Give me full quotes about oral administration","Show full responses about radiation delay","Give me full quotes about patient assistance programs","Show full responses about the Indigo trial","Show full quotes about overall survival"],
    "💰 Cost & Access": ["How do HCPs navigate insurance denials?","What patient assistance programs are mentioned?","How do doctors handle prior authorization?","What approaches do HCPs use for reimbursement?","How do community HCPs handle cost barriers?","What is the role of the rep in cost navigation?"],
    "🏥 Patient Profile": ["What patient profile is ideal for prescribing?","Which patients are most appropriate candidates?","What tumor characteristics matter most?","Do younger patients get prescribed more?","What performance status is needed?","What are the key eligibility factors?"],
    "📋 Clinical Endpoints": ["Do HCPs prefer PFS or OS?","What clinical endpoint matters most?","How many mentioned PFS vs OS?","Is OS or PFS more discussed in academic settings?"],
}

# ── APP ───────────────────────────────────────────────────────────────────────
st.markdown("""<div style="margin-bottom:20px">
    <div style="font-family:'IBM Plex Mono',monospace;font-size:10px;color:#3b6ef7;letter-spacing:2px;margin-bottom:6px">HCP QUAL INSIGHT ENGINE · UPLOAD YOUR DATA TO BEGIN</div>
    <h1 style="font-size:30px;margin:0;color:#e2ecf8;font-weight:600">HCP Insight Engine</h1>
</div>""", unsafe_allow_html=True)

# ── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="sec-lbl">📂 UPLOAD DATA</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:10px;line-height:1.6">Upload your qual Excel.<br><span style="color:#8a9ab5">Required:</span> verbatim/transcript column<br><span style="color:#8a9ab5">Optional:</span> Setting · Specialty · Target</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader("Upload Excel", type=["xlsx","xls"], label_visibility="collapsed")
    st.markdown("<hr>", unsafe_allow_html=True)

    if not uploaded:
        st.markdown('<div style="text-align:center;padding:20px;color:#4a6080;font-size:12px">Upload an Excel file above to begin. All analysis uses your data only.</div>', unsafe_allow_html=True)
        st.stop()

    full_df, err = load_excel(uploaded)
    if err or full_df is None:
        st.error(f"⚠️ {err}")
        st.markdown('<div style="font-size:11px;color:#4a6080;margin-top:8px">Tips:<br>· Needs verbatim text column (>100 chars)<br>· Metadata cols should have few unique values<br>· Re-save as .xlsx if issues persist</div>', unsafe_allow_html=True)
        st.stop()

    TOTAL = len(full_df)
    st.markdown(f'<div style="background:#0a1f14;border:1px solid #10b98133;border-radius:10px;padding:12px 16px;margin-bottom:14px"><div style="font-size:11px;font-weight:600;color:#34d399;margin-bottom:3px">✅ Data loaded</div><div style="font-size:11px;color:#4a6080">{TOTAL} respondents · {uploaded.name}</div></div>', unsafe_allow_html=True)

    st.markdown('<div class="sec-lbl">SEGMENT FILTERS</div>', unsafe_allow_html=True)
    settings    = ["All"] + sorted(full_df["setting"].dropna().unique().tolist())
    specialties = ["All"] + sorted(full_df["specialty"].dropna().unique().tolist())
    targets     = ["All"] + sorted(full_df["target"].dropna().unique().tolist())
    fs  = st.selectbox("Practice Setting", settings)
    fsp = st.selectbox("Specialty",        specialties)
    ft  = st.selectbox("Target Type",      targets)
    fdf = full_df.copy()
    if fs  != "All": fdf = fdf[fdf["setting"]  ==fs]
    if fsp != "All": fdf = fdf[fdf["specialty"]==fsp]
    if ft  != "All": fdf = fdf[fdf["target"]   ==ft]
    fn=len(fdf); sc=fdf["setting"].value_counts().to_dict()
    sh=" · ".join(f'<span style="color:#8a9ab5">{k}:{v}</span>' for k,v in sc.items())
    st.markdown(f'<div class="card" style="text-align:center;margin-top:6px"><div class="stat-num">{fn}</div><div class="stat-lbl">in view</div><div style="font-size:11px;margin-top:6px">{sh}</div></div>', unsafe_allow_html=True)
    st.markdown("<hr>", unsafe_allow_html=True)

    # ── THREE-TAB SIDEBAR ─────────────────────────────────────────────────────
    side_tab1, side_tab2, side_tab3 = st.tabs(["📊 Dashboard", "❓ Questions", "🎭 Sentiment"])

    with side_tab1:
        st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:10px;margin-top:8px;line-height:1.6">Click any theme for the full dashboard — executive summary, segment breakdown, quotes.</div>', unsafe_allow_html=True)
        for theme_name in list(THEMES.keys()):
            if st.button(f"● {theme_name}", key=f"dash_{theme_name}"):
                st.session_state["dashboard_theme"] = theme_name
                st.session_state["q"] = ""
                st.session_state["sent_theme"] = ""
                st.session_state["sent_overall"] = False
                st.rerun()

    with side_tab2:
        st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:8px;margin-top:8px">Click any question to run instantly</div>', unsafe_allow_html=True)
        for cat, qs in QB.items():
            with st.expander(cat, expanded=False):
                for q in qs:
                    if st.button(q, key=f"qb_{q[:28]}"):
                        st.session_state["q"] = q
                        st.session_state["dashboard_theme"] = ""
                        st.session_state["sent_theme"] = ""
                        st.session_state["sent_overall"] = False
                        st.rerun()

    with side_tab3:
        st.markdown('<div style="font-size:11px;color:#4a6080;margin-bottom:10px;margin-top:8px;line-height:1.6">Sentiment analysis — stats and NSS only. All quotes downloadable as Excel.</div>', unsafe_allow_html=True)
        # Overall view button
        if st.button("🌐  All themes — overview", key="sent_overall_btn"):
            st.session_state["sent_overall"] = True
            st.session_state["sent_theme"] = ""
            st.session_state["dashboard_theme"] = ""
            st.session_state["q"] = ""
            st.rerun()
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div style="font-size:10px;color:#4a6080;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">By theme</div>', unsafe_allow_html=True)
        for theme_name in list(THEMES.keys()):
            if st.button(f"● {theme_name}", key=f"sent_{theme_name}"):
                st.session_state["sent_theme"] = theme_name
                st.session_state["sent_overall"] = False
                st.session_state["dashboard_theme"] = ""
                st.session_state["q"] = ""
                st.rerun()

# ── MAIN ──────────────────────────────────────────────────────────────────────
if "q" not in st.session_state: st.session_state["q"] = ""
if "dashboard_theme" not in st.session_state: st.session_state["dashboard_theme"] = ""
if "sent_theme" not in st.session_state: st.session_state["sent_theme"] = ""
if "sent_overall" not in st.session_state: st.session_state["sent_overall"] = False

def back_btn(label="← Back"):
    if st.button(label, key="back_btn"):
        st.session_state["dashboard_theme"] = ""
        st.session_state["sent_theme"] = ""
        st.session_state["sent_overall"] = False
        st.session_state["q"] = ""
        st.rerun()

# ── SENTIMENT OVERALL MODE ────────────────────────────────────────────────────
if st.session_state.get("sent_overall"):
    back_btn("← Back to search")
    render_overall_sentiment(full_df)
    st.stop()

# ── SENTIMENT THEME MODE ──────────────────────────────────────────────────────
if st.session_state.get("sent_theme") and not st.session_state.get("q"):
    back_btn("← Back to search")
    render_theme_sentiment(st.session_state["sent_theme"], full_df)
    st.stop()

# ── DASHBOARD MODE ────────────────────────────────────────────────────────────
if st.session_state.get("dashboard_theme") and not st.session_state.get("q"):
    theme_selected = st.session_state["dashboard_theme"]
    back_btn("← Back to search")
    render_dashboard(theme_selected, full_df)
    st.stop()

# Search label + colour legend row
st.markdown("""
<div style="margin-bottom:8px">
    <div style="font-size:12px;color:#8a9ab5;margin-bottom:6px;font-weight:500">
        🔍 <b style="color:#e2ecf8">Ask a question</b>
        <span style="color:#4a6080;font-weight:400"> — type freely below, or pick from the sidebar browser</span>
    </div>
</div>""", unsafe_allow_html=True)

query = st.text_input(
    label="Ask a question about your data",
    label_visibility="collapsed",
    placeholder='e.g. "How many HCPs mentioned PFS?" · "Community vs academic on cost" · "Was PFS a standalone driver?"',
    value=st.session_state.get("q",""),
    key="main"
)

# Quick-type suggestion chips (clickable, distinct from sidebar browser)
st.markdown('<div style="font-size:10px;color:#4a6080;margin:8px 0 4px;letter-spacing:1px;text-transform:uppercase">Quick questions — click to run:</div>', unsafe_allow_html=True)
chip_cols = st.columns(4)
CHIPS = [
    "How many HCPs mentioned PFS?",
    "Community vs academic on PFS",
    "What other drivers were tagged to PFS?",
    "Was PFS a standalone driver?",
    "Did PFS and quality of life go together?",
    "How many discussed cost or insurance?",
    "Show full responses about cost and insurance",
    "What are the main barriers to prescribing?",
]
for i, chip in enumerate(CHIPS):
    if chip_cols[i % 4].button(chip, key=f"chip_{i}"):
        st.session_state["q"] = chip
        st.rerun()

# Colour legend
st.markdown("""<div style="display:flex;flex-wrap:wrap;gap:8px;margin:12px 0 4px;align-items:center">
    <span style="font-size:10px;color:#4a6080;margin-right:2px">Highlight key:</span>
    <span class="hl-ins" style="font-style:normal">insurance/cost</span>
    <span class="hl-pfs" style="font-style:normal">PFS</span>
    <span class="hl-os" style="font-style:normal">OS</span>
    <span class="hl-saf" style="font-style:normal">safety</span>
    <span class="hl-sz" style="font-style:normal">seizures</span>
    <span class="hl-qol" style="font-style:normal">quality of life</span>
    <span class="hl-oral" style="font-style:normal">oral</span>
    <span class="hl-trial" style="font-style:normal">trial data</span>
    <span class="hl-nccn" style="font-style:normal">NCCN/guidelines</span>
</div>""", unsafe_allow_html=True)

if query and query.strip():
    with st.spinner(""):
        r = answer(query, fdf, full_df)

    st.markdown("<hr>", unsafe_allow_html=True)
    bh = f'<span class="ibadge">{r["intent"]}</span>'
    for t in r["topics"]: bh += f'<span class="tbadge">{t}</span>'
    if fs != "All": bh += f'<span class="ibadge" style="color:#fbbf24;border-color:rgba(251,191,36,.3);background:rgba(251,191,36,.06)">{fs}</span>'
    st.markdown(bh+"<br><br>", unsafe_allow_html=True)

    st.markdown(f"""<div class="card">
        <div class="sec-lbl">INSIGHT SUMMARY</div>
        <div style="font-size:17px;color:#e2ecf8;line-height:1.75">{r["summary"]}</div>
        <div style="font-size:11px;color:#4a6080;margin-top:8px">All counts and quotes drawn directly from uploaded data ({TOTAL} respondents). No inference or generation.</div>
    </div>""", unsafe_allow_html=True)

    T=r["T"]

    # ── COMPARISON ────────────────────────────────────────────────────────
    if r["is_comp"] and r["comp"]:
        cp=r["comp"]; va,vb=cp["va"],cp["vb"]; na,nb=cp["na"],cp["nb"]
        st.markdown(f'<div style="display:flex;gap:12px;margin-bottom:16px;align-items:center"><span class="seg-a">🟡 {va} (n={na})</span><span style="color:#4a6080">vs</span><span class="seg-b">🔵 {vb} (n={nb})</span></div>', unsafe_allow_html=True)
        st.markdown('<div class="card"><div class="sec-lbl">THEME COMPARISON</div>', unsafe_allow_html=True)
        for row in cp["rows"]:
            t=row["Theme"]; pa,pb,d=row["_pa"],row["_pb"],row["_d"]; color=TC.get(t,"#4a6080")
            dh=f'<span class="diff-pos">▲{d}pp</span>' if d>0 else (f'<span class="diff-neg">▼{abs(d)}pp</span>' if d<0 else '<span style="color:#4a6080">═</span>')
            st.markdown(f"""<div style="margin-bottom:12px">
                <div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:3px"><span style="color:#e2ecf8;font-weight:500">{t}</span><span>{dh}</span></div>
                <div style="display:flex;gap:6px;align-items:center;margin-bottom:2px">
                    <span style="font-size:10px;color:#fbbf24;min-width:65px">{va}</span><div class="bar-trk" style="flex:1"><div class="bar-fill" style="width:{pa}%;background:#f59e0b"></div></div><span style="font-family:'IBM Plex Mono',monospace;font-size:11px;color:#fbbf24;min-width:38px;text-align:right">{pa}%</span>
                </div>
                <div style="display:flex;gap:6px;align-items:center">
                    <span style="font-size:10px;color:#93c5fd;min-width:65px">{vb}</span><div class="bar-trk" style="flex:1"><div class="bar-fill" style="width:{pb}%;background:#3b6ef7"></div></div><span style="font-family:'IBM Plex Mono',monospace;font-size:11px;color:#93c5fd;min-width:38px;text-align:right">{pb}%</span>
                </div>
            </div>""", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
        c1,c2=st.columns(2)
        for col,df_s,val,color in [(c1,cp["dfa"],va,"#fbbf24"),(c2,cp["dfb"],vb,"#93c5fd")]:
            with col:
                st.markdown(f'<div class="sec-lbl" style="color:{color}">{"🟡" if color=="#fbbf24" else "🔵"} {val} FULL RESPONSES</div>', unsafe_allow_html=True)
                f=cp.get("focus",[])
                filt=df_s[df_s["text_lower"].apply(lambda x: any(p in x for p in f) if f else True)]
                for _,row in filt.head(4).iterrows(): quote_card(row)
        exp=[{"Theme":row["Theme"],f"{va}(n={na})":row.get(f"{va}(n={na})",""),f"{vb}(n={nb})":row.get(f"{vb}(n={nb})",""),"Diff":row["D"]} for row in cp["rows"]]
        st.download_button("⬇ Download comparison CSV", pd.DataFrame(exp).to_csv(index=False).encode(),"comparison.csv","text/csv")

    # ── CO-OCCURRENCE ─────────────────────────────────────────────────────
    elif r["intent"]=="co_occur" and r.get("co"):
        co=r["co"]; ta,tb=co["ta"],co["tb"]; ca=TC.get(ta,"#3b6ef7"); cb=TC.get(tb,"#f59e0b")
        c1,c2,c3=st.columns(3)
        c1.markdown(f'<div class="card" style="text-align:center;border-left:3px solid {ca}"><div class="stat-num" style="color:{ca}">{co["oa"]}</div><div class="stat-lbl">{ta} only</div></div>', unsafe_allow_html=True)
        c2.markdown(f'<div class="card" style="text-align:center;border:2px solid #10b981"><div class="stat-num" style="color:#10b981">{co["nb"]}</div><div class="stat-lbl">BOTH ({co["pct"]}%)</div></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="card" style="text-align:center;border-left:3px solid {cb}"><div class="stat-num" style="color:{cb}">{co["ob"]}</div><div class="stat-lbl">{tb} only</div></div>', unsafe_allow_html=True)
        if co["rows"]:
            st.markdown(f'<div class="card"><div class="sec-lbl" style="color:#10b981">RESPONDENTS MENTIONING BOTH — FULL RESPONSES WITH ALL THEMES HIGHLIGHTED</div><div style="font-size:11px;color:#4a6080;margin-bottom:12px">Both <b style="color:{ca}">{ta}</b> and <b style="color:{cb}">{tb}</b> highlighted. Full text, no truncation.</div>', unsafe_allow_html=True)
            for row in co["rows"][:5]: quote_card(row, [ta,tb])
            st.markdown('</div>', unsafe_allow_html=True)
            rows=[{"ID":rec["id"],"Setting":rec["setting"],"Target":rec.get("target",""),"Specialty":rec["specialty"],"Full Response":rec["text"]} for rec in co["rows"]]
            st.download_button("⬇ Download co-occurrence CSV", pd.DataFrame(rows).to_csv(index=False).encode(),"co_occurrence.csv","text/csv")

    # ── CLUSTER ───────────────────────────────────────────────────────────
    elif r["intent"]=="cluster" and r.get("cl"):
        cl=r["cl"]; vc={"complex":"#ef4444","moderate":"#f59e0b","standalone":"#10b981"}.get(cl["verdict"],"#4a6080")
        c1,c2,c3=st.columns(3)
        c1.markdown(f'<div class="card" style="text-align:center"><div class="stat-num">{cl["n"]}</div><div class="stat-lbl">mentioned {cl["anchor"]}</div></div>', unsafe_allow_html=True)
        c2.markdown(f'<div class="card" style="text-align:center"><div class="stat-num">{cl["avg"]}</div><div class="stat-lbl">avg other themes</div></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="card" style="text-align:center"><div class="stat-num" style="color:{vc};font-size:22px">{cl["verdict"].upper()}</div><div class="stat-lbl">driver type</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="card"><div class="sec-lbl">THEMES THAT TRAVEL WITH {cl["anchor"].upper()} — FROM UPLOADED DATA ONLY</div>', unsafe_allow_html=True)
        for t,cnt in cl["clusters"]:
            color=TC.get(t,"#4a6080"); pct=round(cnt/cl["n"]*100) if cl["n"] else 0
            st.markdown(f'<div style="margin-bottom:6px"><div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:3px"><span style="color:#e2ecf8">{t}</span><span style="font-family:\'IBM Plex Mono\',monospace;color:{color}">{cnt} ({pct}% of group)</span></div><div class="bar-trk"><div class="bar-fill" style="width:{pct}%;background:{color}"></div></div></div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="sec-lbl" style="margin-top:14px">FULL RESPONSES FROM {cl["anchor"].upper()} GROUP</div>', unsafe_allow_html=True)
        for _,row in cl["adf"].head(5).iterrows(): quote_card(row,[cl["anchor"]])
        exp=[{"Theme":t,"Count":cnt,"% of group":f"{round(cnt/cl['n']*100) if cl['n'] else 0}%"} for t,cnt in cl["clusters"]]
        st.download_button("⬇ Download cluster CSV", pd.DataFrame(exp).to_csv(index=False).encode(),"cluster.csv","text/csv")

    # ── COMPLEXITY ────────────────────────────────────────────────────────
    elif r["intent"]=="complexity" and r.get("cx"):
        cx=r["cx"]; vc_map={"standalone":("#10b981","🟢 STANDALONE","Mentioned independently, not bundled with many other themes."),"complex":("#ef4444","🔴 COMPLEX / ENTANGLED","Almost always appeared alongside 4+ other themes."),"mixed":("#f59e0b","🟡 MIXED","Sometimes standalone, sometimes part of a broader cluster.")}
        vc,vt,vd=vc_map.get(cx.get("verdict","mixed"),("#4a6080","MIXED",""))
        st.markdown(f'<div class="card" style="border-left:4px solid {vc}"><div style="font-size:13px;font-weight:700;color:{vc};margin-bottom:6px">{vt}</div><div style="font-size:13px;color:#c8d4e8;line-height:1.6">{vd}</div><div style="font-size:11px;color:#4a6080;margin-top:8px">Based on {cx["n"]} respondents — from uploaded data only.</div></div>', unsafe_allow_html=True)
        c1,c2,c3=st.columns(3)
        c1.markdown(f'<div class="card" style="text-align:center;border-left:3px solid #10b981"><div class="stat-num" style="color:#10b981">{cx["ns"]}</div><div class="stat-lbl">Standalone (0–1 others)</div></div>', unsafe_allow_html=True)
        c2.markdown(f'<div class="card" style="text-align:center;border-left:3px solid #f59e0b"><div class="stat-num" style="color:#f59e0b">{cx["nm"]}</div><div class="stat-lbl">Moderate (2–3 others)</div></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="card" style="text-align:center;border-left:3px solid #ef4444"><div class="stat-num" style="color:#ef4444">{cx["nc"]}</div><div class="stat-lbl">Complex (4+ others)</div></div>', unsafe_allow_html=True)
        if cx.get("top_co"):
            st.markdown(f'<div class="card"><div class="sec-lbl">THEMES MOST MENTIONED ALONGSIDE {cx["anchor"].upper()}</div>', unsafe_allow_html=True)
            for t,cnt in cx["top_co"].items(): bar_html(t,cnt,cx["n"],TC.get(t,"#4a6080"))
            st.markdown('</div>', unsafe_allow_html=True)
        if cx["solo"]:
            st.markdown('<div class="sec-lbl" style="color:#10b981;margin-top:14px">🟢 STANDALONE — FULL RESPONSES</div>', unsafe_allow_html=True)
            for rec in cx["solo"][:3]: quote_card(rec,[cx["anchor"]])
        if cx["comp"]:
            st.markdown('<div class="sec-lbl" style="color:#ef4444;margin-top:14px">🔴 COMPLEX MENTIONS — FULL RESPONSES</div>', unsafe_allow_html=True)
            for rec in cx["comp"][:3]: quote_card(rec,[cx["anchor"]]+rec.get("others",[])[:4])
        rows=[]
        for cat,recs in [("standalone",cx["solo"]),("moderate",cx["mod"]),("complex",cx["comp"])]:
            for rec in recs: rows.append({"ID":rec["id"],"Setting":rec["setting"],"Specialty":rec["specialty"],"Type":cat,"N other themes":rec["n_other"],"Other themes":", ".join(rec.get("others",[])[:5]),"Full Response":rec["text"]})
        st.download_button("⬇ Download complexity CSV", pd.DataFrame(rows).to_csv(index=False).encode(),"complexity.csv","text/csv")

    # ── STANDARD ─────────────────────────────────────────────────────────
    else:
        l,ri=st.columns([1,1])
        with l:
            cc1,cc2=st.columns(2)
            cc1.markdown(f'<div class="card" style="text-align:center"><div class="stat-num">{r["n"]}</div><div class="stat-lbl">matched</div></div>', unsafe_allow_html=True)
            cc2.markdown(f'<div class="card" style="text-align:center"><div class="stat-num">{round(r["n"]/T*100) if T else 0}%</div><div class="stat-lbl">of total</div></div>', unsafe_allow_html=True)
            for title,data in r["chart"].items():
                if data:
                    st.markdown(f'<div class="card"><div class="sec-lbl">{title}</div>', unsafe_allow_html=True)
                    for lbl,cnt in list(data.items())[:12]: bar_html(lbl,int(cnt),T,TC.get(lbl,"#3b6ef7"))
                    st.markdown('</div>', unsafe_allow_html=True)
        with ri:
            n_rows=len(r["rows"])
            st.markdown(f'<div class="sec-lbl">FULL RESPONSES — ALL THEMES HIGHLIGHTED</div>', unsafe_allow_html=True)
            st.markdown(f'<div style="font-size:11px;color:#4a6080;margin-bottom:12px">Showing {min(n_rows,6)} of {n_rows} matched respondents. Full text shown — keywords colour-coded above.</div>', unsafe_allow_html=True)
            if r["rows"]:
                for rec in r["rows"][:6]: quote_card(rec, r["topics"] if r["topics"] else None)
            else:
                st.markdown('<div style="color:#4a6080;font-size:13px">No matches. Try broader keywords.</div>', unsafe_allow_html=True)
        if r.get("export"):
            st.markdown("<hr>", unsafe_allow_html=True)
            dfe=pd.DataFrame(r["export"])
            preview_cols=[c for c in ["ID","Setting","Specialty","Target"] if c in dfe.columns]
            st.dataframe(dfe[preview_cols].head(20) if preview_cols else dfe.head(20),hide_index=True,use_container_width=True)
            st.download_button("⬇ Download full responses CSV", dfe.to_csv(index=False).encode(),f"responses_{query[:20].replace(' ','_')}.csv","text/csv")

else:
    # Landing
    st.markdown('<div style="text-align:center;margin-top:60px"><div style="font-size:48px;margin-bottom:16px">📂</div><div style="font-size:16px;color:#4a6080">Upload your Excel file in the sidebar to begin</div><div style="font-size:13px;color:#2a3a55;margin-top:8px">Then click any question from the sidebar browser or type your own</div></div>', unsafe_allow_html=True)
